#!/usr/bin/env python3
"""
App信息查询工具 - Web版
通过包名查找App的名称和下载地址
"""

import csv
import io
import json
import math
import os
import sys
import re
import struct
import time
import uuid
import zlib
import zipfile
import concurrent.futures
import random
import threading
import urllib.parse

from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import requests
from bs4 import BeautifulSoup
try:
    from PIL import Image as PILImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

try:
    from cryptography.hazmat.primitives.serialization import pkcs7 as crypto_pkcs7
    from cryptography.hazmat.primitives import hashes as crypto_hashes
    HAS_CRYPTOGRAPHY = True
except ImportError:
    HAS_CRYPTOGRAPHY = False


def _lower_process_priority():
    """把本进程降为低优先级，大量并发查询时不会挤占其他程序的 CPU。
    - Windows: IDLE_PRIORITY_CLASS（系统完全空闲时才跑，前台用 CPU 时全让出）
    - macOS / Linux: nice +10
    设计目标：批量查询时用户照常打字/看视频/做其他事，工具在后台"偷 CPU"跑，
    用户一抢 CPU 就立刻让出，完全不影响前台体验。"""
    try:
        if os.name == "nt":
            import ctypes
            # IDLE_PRIORITY_CLASS = 0x00000040 比 BELOW_NORMAL (0x00004000) 更低
            # 只在系统空闲时获得 CPU 时间，前台程序要用 CPU 时本进程让出
            IDLE_PRIORITY_CLASS = 0x00000040
            kernel32 = ctypes.windll.kernel32
            kernel32.SetPriorityClass(kernel32.GetCurrentProcess(),
                                      IDLE_PRIORITY_CLASS)
        else:
            if hasattr(os, "nice"):
                os.nice(10)
    except Exception:
        pass


_lower_process_priority()


# ============================================================
# 性能优化：共享 HTTP 会话 + 选用最快的 HTML 解析器
# ============================================================

# 1) lxml 解析器比内置的 html.parser 快 3~5 倍，CPU 占用明显低
#    没装 lxml 时自动回退到 html.parser，行为完全一致
try:
    import lxml  # noqa: F401
    _BS_PARSER = "lxml"
except Exception:
    _BS_PARSER = "html.parser"


# 2) 全局共享的 requests.Session：
#    - 对同一台商店的第 2 次起请求复用 TCP/TLS 连接（省 50~200ms/次）
#    - 每个 host 最多 50 个池连接，足够支撑并发 20 个外层 worker
#    - 自带 HTTPS/urllib3 底层线程安全，GET/HEAD 并发无需加锁
from requests.adapters import HTTPAdapter
_HTTP = requests.Session()
_HTTP_ADAPTER = HTTPAdapter(pool_connections=32, pool_maxsize=64, max_retries=0)
_HTTP.mount("http://",  _HTTP_ADAPTER)
_HTTP.mount("https://", _HTTP_ADAPTER)


# 3) 模块级共享线程池：供 query_single 内部的商店并发查询复用
#    - 避免每次查询都 new 一个 ThreadPoolExecutor（批量 50 条 = 50 次池创建/销毁）
#    - 工作线程数足够覆盖：6 个主商店 + apple + flyme + 3 个搜索引擎 + 名称回查 = 16
#    - daemon 线程，进程退出时自动回收
_QUERY_POOL = concurrent.futures.ThreadPoolExecutor(
    max_workers=32, thread_name_prefix="query_io"
)


def _parallel_map(func, items, timeout=None):
    """并发执行 func(item) 返回 {item: result}。异常 / 超时记为 None。
    用于批量场景下并行补齐（如 appchina APK 直链）。"""
    if not items:
        return {}
    results = {}
    futures = {_QUERY_POOL.submit(func, it): it for it in items}
    try:
        for f in concurrent.futures.as_completed(futures, timeout=timeout):
            it = futures[f]
            try:
                results[it] = f.result(timeout=0)
            except Exception:
                results[it] = None
    except concurrent.futures.TimeoutError:
        # 超时时把未完成的记为 None
        for f, it in futures.items():
            if it not in results:
                results[it] = None
    return results

# ── 连接会话自愈 ─────────────────────────────────────────────────────
# 场景：服务跑久了后，上游商店（小米/七麦/搜狗）可能给 Session 下发限流
# cookie 或把 IP 标记为异常，此后所有搜索请求都会空返。
# 应对：侦测到连续 N 次查询"完全无结果"时，自动销毁旧 Session + 重建连接池。
# 用户无感，无需重启服务。
_SESSION_HEALTH_LOCK       = threading.Lock()
_SESSION_HEALTH_EMPTY_STREAK = 0                 # 连续空返回计数
_SESSION_LAST_RESET_AT     = 0.0                 # 上次重置时间戳
_SESSION_RESET_THRESHOLD   = 5                   # 连续空 N 次触发重置
_SESSION_RESET_COOLDOWN    = 60.0                # 两次重置间隔下限（秒）


def _reset_http_session():
    """销毁旧的 _HTTP Session 并新建一个，清空 cookie 和 TCP 连接池。
    线程安全：仅在 _SESSION_HEALTH_LOCK 下调用。"""
    global _HTTP, _HTTP_ADAPTER, _SESSION_HEALTH_EMPTY_STREAK, _SESSION_LAST_RESET_AT
    try:
        _HTTP.close()
    except Exception:
        pass
    new_session = requests.Session()
    new_adapter = HTTPAdapter(pool_connections=32, pool_maxsize=64, max_retries=0)
    new_session.mount("http://",  new_adapter)
    new_session.mount("https://", new_adapter)
    _HTTP = new_session
    _HTTP_ADAPTER = new_adapter
    _SESSION_HEALTH_EMPTY_STREAK = 0
    _SESSION_LAST_RESET_AT = time.time()


def _note_query_result(had_real_result):
    """每完成一条查询调用一次。连续 N 次空返时触发 Session 重置。
    返回 True 表示本次调用刚刚触发了重置（给调用方用来发通知/再查一次）。"""
    global _SESSION_HEALTH_EMPTY_STREAK
    with _SESSION_HEALTH_LOCK:
        if had_real_result:
            _SESSION_HEALTH_EMPTY_STREAK = 0
            return False
        _SESSION_HEALTH_EMPTY_STREAK += 1
        if _SESSION_HEALTH_EMPTY_STREAK < _SESSION_RESET_THRESHOLD:
            return False
        # 防抖：刚重置完不久又空返，多半是网络本身有问题，不要反复重建
        if time.time() - _SESSION_LAST_RESET_AT < _SESSION_RESET_COOLDOWN:
            return False
        _reset_http_session()
        return True


# PyInstaller 打包后资源路径处理
def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, relative_path)


app = Flask(
    __name__,
    template_folder=resource_path('templates'),
    static_folder=resource_path('static'),
)
app.config['TEMPLATES_AUTO_RELOAD'] = True


@app.before_request
def _lan_access_gate():
    """LAN 访问门禁 + 统计：
    - 本机请求（127.*）永远允许
    - LAN 关闭时：非本机请求返回 403
    - LAN 开启时：记录来源 IP 和请求次数到 _LAN_STATS
    """
    client_ip = request.remote_addr or ""
    if _is_local_request_ip(client_ip):
        return  # 本机请求放行，不计入 LAN 统计

    if not _LAN_ENABLED:
        return (jsonify({
            "error": "lan_access_disabled",
            "message": "本机管理员未开启 LAN 共享，其他设备暂时无法访问"
        }), 403)

    # 黑名单检查：被管理员屏蔽的 IP 一律拒绝
    if client_ip in _LAN_BLOCKED_IPS:
        return (jsonify({
            "error": "blocked",
            "message": "当前设备已被管理员屏蔽，无法使用本工具"
        }), 403)

    # LAN 开启：记录设备最后活跃时间 + UA；但"查询次数"只统计真正的查询 API，
    # 不把 JS/CSS/图标/SSE 心跳/stats 轮询等累加进去——否则同一次用户点击
    # "查询"会显示成几十次请求，完全脱离直觉。
    QUERY_PATHS = ("/api/query", "/api/start_job", "/api/retry")
    is_query = request.path in QUERY_PATHS
    # 只在访问"真实功能页面"时才建立设备记录（排除 favicon、图标下载等被动请求）
    IGNORE_PATHS = ("/api/lan_info", "/api/lan_stats", "/favicon.ico")
    if request.path in IGNORE_PATHS:
        return

    now = time.time()
    is_new_device = False
    with _LAN_STATS_LOCK:
        dev = _LAN_STATS["devices"].get(client_ip)
        if dev is None:
            is_new_device = True
            dev = {
                "first_seen": now, "last_seen": now,
                "query_count": 0, "last_path": "",
                "ua": (request.headers.get("User-Agent", "") or "")[:120],
                "mac": "",       # 稍后填
                "hostname": "",  # 稍后填
            }
            _LAN_STATS["devices"][client_ip] = dev
        dev["last_seen"] = now
        dev["last_path"] = request.path
        if is_query:
            dev["query_count"] += 1
            _LAN_STATS["total_requests"] += 1
    # 锁外做：MAC / hostname 探测可能耗时（1-2s），不阻塞主请求
    if is_new_device:
        def _identify(ip, device_entry):
            mac = _get_mac_for_ip(ip)
            host = _get_hostname_for_ip(ip, timeout=1.0)
            with _LAN_STATS_LOCK:
                if device_entry:
                    device_entry["mac"] = mac or ""
                    device_entry["hostname"] = host or ""
        threading.Thread(target=_identify, args=(client_ip, dev), daemon=True).start()

# ========== 后台 Job 管理 ==========
# 结构：{job_id: {status, events, tasks, worker_params, results, created_at}}
JOBS = {}
JOBS_LOCK = threading.Lock()
_JOBS_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".jobs_cache.json")


_LAN_SETTINGS_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), ".lan_settings.json"
)
# LAN 访问开关：默认关闭（出于安全 / 隐私考虑，局域网其他设备默认不能访问本服务）
_LAN_ENABLED = False
# LAN 访问统计：记录连接设备及其请求次数
_LAN_STATS = {"devices": {}, "total_requests": 0}
_LAN_STATS_LOCK = threading.Lock()
# 设备管理：备注（管理员给 IP 起的友好名字）+ 黑名单（被屏蔽的 IP，返回 403）
# 按 IP 为 key 持久化，即便 .app 重启也保留
_LAN_DEVICE_NOTES = {}   # {"192.168.0.102": "办公室 Windows"}
_LAN_BLOCKED_IPS = set() # {"192.168.0.99"}

# 服务端共享查询历史：所有客户端（本机 / LAN 访客）共享，LAN 场景下手机等新设备也能看到
_HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".history.json")
_HISTORY_MAX = 30           # 最多保留多少条
_HISTORY_LOCK = threading.Lock()
_HISTORY = []               # 内存中的列表，顺序：新 → 旧


def _load_lan_settings():
    """启动时从磁盘读取 LAN 开关设置 + 设备备注 + 黑名单"""
    global _LAN_ENABLED, _LAN_DEVICE_NOTES, _LAN_BLOCKED_IPS
    try:
        if os.path.exists(_LAN_SETTINGS_FILE):
            with open(_LAN_SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                _LAN_ENABLED = bool(data.get("enabled", False))
                notes = data.get("device_notes") or {}
                if isinstance(notes, dict):
                    _LAN_DEVICE_NOTES = {str(k): str(v) for k, v in notes.items() if v}
                blocked = data.get("blocked_ips") or []
                if isinstance(blocked, list):
                    _LAN_BLOCKED_IPS = set(str(x) for x in blocked)
    except Exception:
        pass


def _save_lan_settings():
    """持久化 LAN 开关 + 设备备注 + 黑名单"""
    try:
        with open(_LAN_SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "enabled": _LAN_ENABLED,
                "device_notes": _LAN_DEVICE_NOTES,
                "blocked_ips": sorted(_LAN_BLOCKED_IPS),
            }, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


_LOCAL_IPS_CACHE = None


def _get_all_local_ips():
    """枚举本机所有 IP 地址。用于请求来源自判定——
    同一台机器上访问自己的 LAN IP（如浏览器在本机打开 http://192.168.x.x:9527）
    从 socket 层面看 remote_addr 也是那个 LAN IP，但实际是本机自身，
    应当放行，不受 LAN 开关管制。"""
    import socket
    import subprocess

    ips = {"127.0.0.1", "0.0.0.0", "::1"}
    # Method 1: gethostbyname_ex
    try:
        _, _, addrs = socket.gethostbyname_ex(socket.gethostname())
        ips.update(addrs)
    except Exception:
        pass
    # Method 2 (Mac): ipconfig getifaddr 逐个网卡
    if sys.platform == "darwin":
        for iface in ["en0", "en1", "en2", "en3", "en4", "en5"]:
            try:
                r = subprocess.run(
                    ["ipconfig", "getifaddr", iface],
                    capture_output=True, text=True, timeout=1.5,
                )
                ip = (r.stdout or "").strip()
                if ip:
                    ips.add(ip)
            except Exception:
                continue
    return ips


def _is_local_request_ip(ip):
    """判断请求来源是否是本机（含本机自己的 LAN IP）。
    不是本机 = 其他设备通过 LAN 访问，要受 LAN 开关管制。"""
    global _LOCAL_IPS_CACHE
    if not ip:
        return True
    if ip.startswith("127.") or ip == "::1" or ip == "0.0.0.0":
        return True
    # 懒加载：首次调用时探测所有本机 IP
    if _LOCAL_IPS_CACHE is None:
        _LOCAL_IPS_CACHE = _get_all_local_ips()
    return ip in _LOCAL_IPS_CACHE


def _get_mac_for_ip(ip):
    """从系统 ARP 缓存读取 IP 对应的 MAC 地址。
    只要客户端和本机在同一 L2 网段（同一 Wi-Fi / 交换机）就能拿到。
    返回空字符串表示没查到（可能 IP 已过期、在别的网段、或 iOS 私密地址）。"""
    import subprocess
    import re
    if not ip or _is_local_request_ip(ip):
        return ""
    try:
        r = subprocess.run(
            ["arp", "-n", ip],
            capture_output=True, text=True, timeout=1.5,
        )
        m = re.search(r"\bat\s+([0-9a-fA-F:]+)\b", r.stdout or "")
        if m:
            mac = m.group(1).lower()
            if mac != "(incomplete)" and len(mac) >= 11:
                return mac
    except Exception:
        pass
    return ""


def _get_hostname_for_ip(ip, timeout=1.0):
    """IP 反查主机名（mDNS / Bonjour / DNS PTR）。
    Mac 和 iOS 在本地网络下常报告 hostname 如 'Rick-iPhone.local'，非常好辨认。
    Android / Windows 默认不广播 hostname，可能返回空。"""
    import socket
    if not ip:
        return ""
    # socket.gethostbyaddr 不支持直接 timeout，开后台线程执行并设置超时
    import threading as _th
    result = {"name": ""}
    def _q():
        try:
            name, _, _ = socket.gethostbyaddr(ip)
            if name and not name.replace(".", "").isdigit():
                result["name"] = name
        except Exception:
            pass
    t = _th.Thread(target=_q, daemon=True)
    t.start()
    t.join(timeout)
    return result["name"]


def _load_history():
    global _HISTORY
    try:
        if os.path.exists(_HISTORY_FILE):
            with open(_HISTORY_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    _HISTORY = data[:_HISTORY_MAX]
    except Exception:
        pass


def _save_history():
    try:
        with open(_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(_HISTORY, f, ensure_ascii=False)
    except Exception:
        pass


def _save_job_result(job_id, results):
    """将完成的 job 结果持久化到磁盘，供服务重启后恢复"""
    try:
        cache = {}
        if os.path.exists(_JOBS_CACHE_FILE):
            with open(_JOBS_CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
        cache[job_id] = {"results": results, "created_at": time.time()}
        # 只保留最近 10 个
        if len(cache) > 10:
            oldest = sorted(cache, key=lambda k: cache[k]["created_at"])
            for k in oldest[:-10]:
                del cache[k]
        with open(_JOBS_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
    except Exception:
        pass


def _load_job_result(job_id):
    """从磁盘加载已完成 job 的结果"""
    try:
        if not os.path.exists(_JOBS_CACHE_FILE):
            return None
        with open(_JOBS_CACHE_FILE, "r", encoding="utf-8") as f:
            cache = json.load(f)
        entry = cache.get(job_id)
        if not entry:
            return None
        # 超过 2 小时视为过期
        if time.time() - entry.get("created_at", 0) > 7200:
            return None
        return entry["results"]
    except Exception:
        return None


def _cleanup_jobs():
    """清除超过 2 小时的旧 job"""
    cutoff = time.time() - 7200
    with JOBS_LOCK:
        expired = [jid for jid, j in JOBS.items() if j['created_at'] < cutoff]
        for jid in expired:
            del JOBS[jid]


def _parse_query_input(req_data):
    """解析查询请求，返回 (tasks, worker_params, meta, est_seconds)"""
    raw_inputs          = req_data.get("package_names", [])
    exact_search        = bool(req_data.get("exact_search", False))
    get_apk_url         = bool(req_data.get("get_apk_url", False))
    apk_url_mode        = req_data.get("apk_url_mode", "single")
    get_sha1            = bool(req_data.get("get_sha1", False))
    get_sha256          = bool(req_data.get("get_sha256", False))
    platform_filter     = req_data.get("platform_filter", "all")
    query_interval_ms   = max(0, int(req_data.get("query_interval_ms", 0)))

    MAX_ITEMS = 10000
    pkg_list, ios_id_list, name_list = [], [], []
    invalid_count = 0
    all_valid = [x.strip() for x in raw_inputs if x.strip()]
    total_raw = len(all_valid)
    skipped_over_limit = max(0, total_raw - MAX_ITEMS)

    for item in all_valid[:MAX_ITEMS]:
        if _IOS_ID_RE.match(item):
            ios_id_list.append(item)
        else:
            cleaned = clean_package_name(item)
            if cleaned and is_package_name(cleaned):
                pkg_list.append(cleaned)
            elif re.search(r'[a-zA-Z\u4e00-\u9fff]', item):
                name_list.append(item)
            else:
                invalid_count += 1

    def dedup(lst):
        seen, out = set(), []
        for x in lst:
            if x not in seen:
                seen.add(x); out.append(x)
        return out

    cleaned_pkgs    = dedup(pkg_list)
    cleaned_ios_ids = dedup(ios_id_list)
    cleaned_names   = dedup(name_list)
    total_valid     = len(cleaned_pkgs) + len(cleaned_ios_ids) + len(cleaned_names)
    deduplicated    = max(0, (total_raw - skipped_over_limit) - total_valid - invalid_count)

    meta = {
        "over_limit":           skipped_over_limit,
        "total_input":          total_raw,
        "invalid_count":        invalid_count,
        "deduplicated":         deduplicated,
        "name_search_ios_only": False,
    }

    tasks = (
        [("pkg",    p) for p in cleaned_pkgs]    +
        [("ios_id", a) for a in cleaned_ios_ids] +
        [("name",   n) for n in cleaned_names]
    )
    total_tasks = len(tasks)
    user_interval_s = query_interval_ms / 1000.0 if query_interval_ms > 0 else 0.0

    # 并发上限与 CPU 核数挂钩：每个外层 worker 内部还会再派生 9 个商店子线程，
    # 外层开太大导致总线程数爆炸、CPU 飙到 80%+。经验值 min(cpu_count, 8)。
    _CPU_CAP = min((os.cpu_count() or 4), 8)
    if total_tasks <= 500:
        BATCH_SIZE  = total_tasks or 1
        WORKERS     = min(_CPU_CAP, max(1, total_tasks))
        BATCH_DELAY = user_interval_s
        est_seconds = int(math.ceil(total_tasks / max(WORKERS, 1)) * 4 + total_tasks * user_interval_s)
    else:
        BATCH_SIZE  = 20
        WORKERS     = _CPU_CAP
        BATCH_DELAY = max(1.0, user_interval_s)
        batches     = math.ceil(total_tasks / BATCH_SIZE)
        est_seconds = int(batches * (7 + BATCH_DELAY))

    worker_params = {
        "exact_search":        exact_search,
        "get_apk_url":         get_apk_url,
        "apk_url_mode":        apk_url_mode,
        "get_sha1":            get_sha1,
        "get_sha256":          get_sha256,
        "platform_filter":     platform_filter,
        "BATCH_SIZE":          BATCH_SIZE,
        "WORKERS":             WORKERS,
        "BATCH_DELAY":         BATCH_DELAY,
        "meta":                meta,
        "est_seconds":         est_seconds,
        "total_tasks":         total_tasks,
    }

    return tasks, worker_params, meta, est_seconds


def _job_worker(job_id):
    """后台线程：执行查询并将事件追加到 JOBS[job_id]['events']"""
    job = JOBS.get(job_id)
    if not job:
        return

    wp          = job['worker_params']
    tasks       = job['tasks']
    total_tasks = wp['total_tasks']
    BATCH_SIZE  = wp['BATCH_SIZE']
    WORKERS     = wp['WORKERS']
    BATCH_DELAY = wp['BATCH_DELAY']
    meta        = wp['meta']
    est_seconds = wp['est_seconds']

    exact_search        = wp['exact_search']
    get_apk_url         = wp['get_apk_url']
    apk_url_mode        = wp['apk_url_mode']
    get_sha1            = wp['get_sha1']
    get_sha256          = wp['get_sha256']
    platform_filter     = wp['platform_filter']

    def push(event):
        job['events'].append(event)

    if total_tasks == 0:
        push({'type': 'complete', 'results': [], **meta})
        job['status'] = 'done'
        job['results'] = []
        return

    push({'type': 'start', 'total': total_tasks, 'estimated_seconds': est_seconds, **meta})

    all_results = []
    seen_keys   = set()
    done_count  = 0

    try:
        for batch_start in range(0, len(tasks), BATCH_SIZE):
            # 用户取消时立即停止
            if job.get('cancelled'):
                break

            batch = tasks[batch_start: batch_start + BATCH_SIZE]

            with concurrent.futures.ThreadPoolExecutor(
                    max_workers=min(WORKERS, len(batch))) as exe:
                future_map = {}
                for task_type, value in batch:
                    if task_type == "pkg":
                        if get_apk_url or get_sha1 or get_sha256:
                            f = exe.submit(query_single_extended, value,
                                           None, get_apk_url,
                                           apk_url_mode, get_sha1, get_sha256)
                        else:
                            f = exe.submit(query_single, value)
                    elif task_type == "ios_id":
                        f = exe.submit(search_apple_by_numid, value)
                    else:
                        if get_apk_url or get_sha1 or get_sha256:
                            f = exe.submit(query_by_name_extended, value,
                                           None, exact_search,
                                           get_apk_url, apk_url_mode, get_sha1, get_sha256)
                        else:
                            f = exe.submit(query_by_name, value, None, exact_search)
                    future_map[f] = (task_type, value)

                for f in concurrent.futures.as_completed(future_map):
                    task_type, value = future_map[f]
                    try:
                        result = f.result()
                    except Exception:
                        result = None

                    rows = _make_fallback_rows(task_type, value, result)

                    # ── 会话自愈侦测 ──
                    # 判断本次"是否拿到真正结果"：rows 里至少一条不是 _mark_incomplete 兜底。
                    had_real = any(not r.get("incomplete") for r in rows)
                    if _note_query_result(had_real):
                        # 刚刚触发 Session 重置 → 仅通知前端，不立即重查
                        # 原先的同步重查会在批量场景下把请求量翻倍、拖慢整体进度；
                        # 补齐阶段会自动把这一批 incomplete 项用新 Session 重跑一遍
                        push({'type': 'session_reset'})

                    new_rows = []
                    for r in rows:
                        plat = r.get("platform", "")
                        if platform_filter == "ios" and plat != "iOS":
                            continue
                        if platform_filter == "android" and plat != "Android":
                            continue
                        k = (r.get("package_name", ""), plat)
                        if k not in seen_keys:
                            seen_keys.add(k)
                            all_results.append(r)
                            new_rows.append(r)

                    done_count += 1
                    # 将本条新增结果随 progress 事件一起推送，实现前端动态展示
                    push({'type': 'progress', 'done': done_count, 'total': total_tasks,
                          'rows': new_rows})

            if BATCH_DELAY > 0 and batch_start + BATCH_SIZE < len(tasks):
                time.sleep(BATCH_DELAY * random.uniform(0.8, 1.2))

        # ── 自动补齐阶段 ──
        # 首轮批量跑完后，对不完整的条目再静默跑一次（超时翻倍），尽量补齐结果。
        # 包括：按包名查询的 Android 空返、按名称搜索的全面空返（会话自愈后需重查）、
        # 批量并发下连接池吃紧导致慢商店（小米/应用宝）挤不进首轮超时窗口的情况。
        if not job.get('cancelled') and all_results:
            # 收集待补齐项：(task_type, value, key_for_dedup)
            retry_items = []
            seen_retry = set()
            for r in all_results:
                if not r.get("incomplete"):
                    continue
                ttype = r.get("_orig_task_type")
                val   = r.get("_orig_value")
                if not ttype or not val:
                    # 兼容没带原任务信息的老行
                    if r.get("platform") == "Android" and r.get("package_name"):
                        ttype, val = "pkg", r["package_name"]
                    else:
                        continue
                dkey = (ttype, val)
                if dkey in seen_retry:
                    continue
                seen_retry.add(dkey)
                retry_items.append((ttype, val))

            if retry_items:
                retry_timeout = HTTP_TIMEOUT * 2
                retry_total = len(retry_items)
                # 通知前端进入补齐阶段，便于展示 "正在补齐 X 个" 指示
                push({'type': 'retry_start', 'retry_total': retry_total})

                def _retry_one(ttype, val):
                    try:
                        if ttype == "pkg":
                            if get_apk_url or get_sha1 or get_sha256:
                                return query_single_extended(
                                    val, None, get_apk_url,
                                    apk_url_mode, get_sha1, get_sha256,
                                    timeout_override=retry_timeout,
                                ) or []
                            return query_single(val, timeout_override=retry_timeout) or []
                        if ttype == "ios_id":
                            r_one = search_apple_by_numid(val)
                            return [r_one] if r_one else []
                        # name
                        if get_apk_url or get_sha1 or get_sha256:
                            return query_by_name_extended(
                                val, None, exact_search,
                                get_apk_url, apk_url_mode, get_sha1, get_sha256) or []
                        return query_by_name(val, None, exact_search) or []
                    except Exception:
                        return []

                retry_done = 0
                with concurrent.futures.ThreadPoolExecutor(
                        max_workers=min(WORKERS, retry_total)) as exe:
                    fut_map = {exe.submit(_retry_one, t, v): (t, v) for t, v in retry_items}
                    for f in concurrent.futures.as_completed(fut_map):
                        if job.get('cancelled'):
                            break
                        ttype, val = fut_map[f]
                        try:
                            fresh_rows = f.result() or []
                        except Exception:
                            fresh_rows = []

                        # 给新行补上原任务标记，便于后续追溯
                        for fr in fresh_rows:
                            fr.setdefault("_orig_task_type", ttype)
                            fr.setdefault("_orig_value", val)

                        replace_rows = []
                        # 把同 _orig_value 的旧不完整行替换掉
                        # 同一输入可能对应多条结果（iOS + Android），按 platform 匹配
                        handled_idx = set()
                        for fr in fresh_rows:
                            if platform_filter == "ios" and fr.get("platform") != "iOS":
                                continue
                            if platform_filter == "android" and fr.get("platform") != "Android":
                                continue
                            found = False
                            for i, old in enumerate(all_results):
                                if i in handled_idx:
                                    continue
                                if (old.get("_orig_value") == val
                                        and old.get("incomplete")
                                        and (old.get("platform") == fr.get("platform")
                                             or old.get("platform") == "未知")):
                                    merged_row = {**old, **fr}
                                    if old.get("icon_url") and not fr.get("icon_url"):
                                        merged_row["icon_url"] = old["icon_url"]
                                    if old.get("category") and not fr.get("category"):
                                        merged_row["category"] = old["category"]
                                    if (old.get("app_name") and old["app_name"] != "未找到"
                                            and (not fr.get("app_name") or fr["app_name"] == "未找到")):
                                        merged_row["app_name"] = old["app_name"]
                                    all_results[i] = merged_row
                                    handled_idx.add(i)
                                    replace_rows.append(merged_row)
                                    found = True
                                    break
                            if not found:
                                # 新平台的补齐结果（例如原本只有 iOS "未找到" 占位，
                                # 补齐时找到了对应 Android 版本）——作为新行追加
                                k = (fr.get("package_name", ""), fr.get("platform", ""))
                                if k not in seen_keys:
                                    seen_keys.add(k)
                                    all_results.append(fr)
                                    replace_rows.append(fr)

                        retry_done += 1
                        push({
                            'type': 'retry_progress',
                            'retry_done': retry_done,
                            'retry_total': retry_total,
                            'rows': replace_rows,
                        })

                push({'type': 'retry_done'})

        push({'type': 'complete', 'results': all_results, **meta})
        job['status'] = 'done'
        job['results'] = all_results
        _save_job_result(job_id, all_results)

    except Exception:
        job['status'] = 'done'
        job['results'] = all_results
        push({'type': 'complete', 'results': all_results, **meta})
        _save_job_result(job_id, all_results)


# ========== 工具函数 ==========

def clean_package_name(name):
    """清理包名：去掉前后空格、不可见字符、特殊符号"""
    name = re.sub(r'[\u200b\u200c\u200d\ufeff\u00a0\r\n\t\u2028\u2029\u200e\u200f\u202a-\u202e]', '', name)
    name = name.strip()
    # 包名只允许字母、数字、点、下划线、连字符
    if name and not re.match(r'^[a-zA-Z0-9._\-]+$', name):
        return ''
    return name


def clean_app_name(name):
    """清理App名称：去掉副标题/介绍部分"""
    for sep in [' - ', '-', '—', '－', ' – ', ':', '：', '|']:
        if sep in name:
            left = name.split(sep)[0].strip()
            # 只在左边部分有实际内容时才截断（避免名字本身就很短的情况）
            if len(left) >= 1:
                name = left
                break
    return name.strip()


def strip_app_suffix(name):
    """去掉商店惯用的冗余后缀，使名称更简洁。
    '豆包app' → '豆包'，'剪映App' → '剪映'，'WhatsApp' 不受影响（前面是字母）。
    """
    if not name:
        return name
    # 去掉末尾的 app/App/APP/apk/APK，前面必须是非英文字母才处理
    cleaned = re.sub(r'(?<![a-zA-Z])[\s\-·_]*[Aa][Pp][Pp]$', '', name).strip()
    cleaned = re.sub(r'(?<![a-zA-Z])[\s\-·_]*[Aa][Pp][Kk]$', '', cleaned).strip()
    # 去掉末尾常见中文后缀
    cleaned = re.sub(r'[\s·\-]*(安卓版|手机版|官方版|正式版|官方)$', '', cleaned).strip()
    return cleaned if cleaned else name


def _names_related(a, b):
    """判断两个 app 名称是否"同一个 App 的变体"。
    场景：同一个包名在不同商店可能返回"豆包"/"豆包app"/"豆包-AI助手"——算相关；
    但如果商店把旧包配成了完全不同的 App（如 com.qiekj.user 返回"三国" vs 另一家"胖乖生活"），
    两个名字无字符重叠，就认定为不相关，避免 pick_best_name 误选短的那个。
    """
    if not a or not b:
        return False
    if a == b:
        return True
    al, bl = a.lower(), b.lower()
    if al in bl or bl in al:
        return True
    sa, sb = set(al), set(bl)
    # 只考虑中日韩汉字 + 字母数字，忽略空白/标点
    def _filter(s):
        return {c for c in s if c.isalnum() or ('\u4e00' <= c <= '\u9fff')}
    sa, sb = _filter(sa), _filter(sb)
    if not sa or not sb:
        return False
    overlap = len(sa & sb) / max(1, min(len(sa), len(sb)))
    return overlap >= 0.5


def pick_best_name(names, primary=None):
    """从多个商店返回的名称中挑选最简洁的一个。
    策略：
    1. 清洗 + 去冗余后缀
    2. 以 `primary` 作为基准名（来自最可信的来源，如按商店优先级选出的 best）。
       未指定 primary 时退回到第一个非空候选。
    3. 只在与基准名"相关"的候选中取最短——防止某家商店把包名配成完全不同的 App
       导致名称被错换（如 com.qiekj.user 被换成"三国"而不是"胖乖生活"）。
    """
    candidates = []
    for n in names:
        if not n or n == "未找到":
            continue
        stripped = strip_app_suffix(clean_app_name(n))
        if stripped and stripped not in candidates:
            candidates.append(stripped)
    if not candidates:
        return ""
    # 基准名优先使用调用方传入的 primary（优先级来源），否则用第一个候选
    if primary:
        primary_clean = strip_app_suffix(clean_app_name(primary))
    else:
        primary_clean = candidates[0]
    if not primary_clean:
        primary_clean = candidates[0]
    related = [primary_clean]
    for c in candidates:
        if c == primary_clean:
            continue
        if _names_related(primary_clean, c) and c not in related:
            related.append(c)
    return min(related, key=len)


def clean_ios_url(url):
    """简化iOS App Store链接"""
    match = re.search(r'(https://apps\.apple\.com/\w+/app/)(.*?)(id\d+)', url)
    if match:
        return f"{match.group(1)}{match.group(3)}"
    return url


def is_package_name(s):
    """判断输入是否为包名格式（含点号的字母数字组合）"""
    return bool(re.match(r'^[a-zA-Z0-9._\-]+$', s)) and '.' in s


def _make_fallback_rows(task_type, value, result):
    """整理查询结果行，无结果时保证七麦兜底。
    标记 _orig_task_type / _orig_value，供会话自愈 + 补齐阶段按原 task_type 正确重查。"""
    if task_type == "pkg":
        rows = result or []
        if not rows:
            rows = [_mark_incomplete({
                "package_name": value, "platform": "Android",
                "app_name": "未找到",
                "download_url": f"https://www.qimai.cn/search/android/search/{urllib.parse.quote(value)}",
                "icon_url": "", "category": "",
                "source": "qimai_hint",
            })]
        for r in rows:
            r.setdefault("_orig_task_type", "pkg")
            r.setdefault("_orig_value", value)
        return rows
    elif task_type == "ios_id":
        if result:
            result.setdefault("_orig_task_type", "ios_id")
            result.setdefault("_orig_value", value)
            return [result]
        return [{
            "package_name": value, "platform": "iOS",
            "app_name": "未找到",
            "download_url": f"https://apps.apple.com/cn/app/id{value}",
            "icon_url": "", "category": "",
            "_orig_task_type": "ios_id", "_orig_value": value,
        }]
    else:  # name search
        rows = result or []
        if not rows:
            rows = [_mark_incomplete({
                "package_name": value, "platform": "未知",
                "app_name": "未找到",
                "download_url": f"https://www.qimai.cn/search/android/search/{urllib.parse.quote(value)}",
                "icon_url": "", "category": "",
                "source": "qimai_hint",
            })]
        for r in rows:
            r.setdefault("_orig_task_type", "name")
            r.setdefault("_orig_value", value)
        return rows


def _is_apk_direct_url(url):
    """判断 URL 是否是「APK 直链」（时效性 CDN 下载链接），而不是稳定的商店详情页。
    这类 URL 不应放到"商店地址"列（会过期，且点开不是商店页面）。"""
    if not url:
        return False
    low = url.lower()
    # 路径以 .apk 结尾
    path = low.split("?", 1)[0]
    if path.endswith(".apk") or ".apk/" in path:
        return True
    # 常见 APK 直链参数（搜狗 / 腾讯 CDN 都有 fsname=xxx.apk）
    if "fsname=" in low and ".apk" in low:
        return True
    # 已知的 APK CDN host
    apk_hosts = ("imtt.dd.qq.com", "downpack.baidu.com", "app.gdown.baidu.com")
    if any(h in low for h in apk_hosts):
        return True
    return False


def _qimai_fallback_url(pkg_or_name):
    """七麦搜索页兜底 URL：对任何未找到 / 信息不完整的结果都可用，稳定且不过期。"""
    if not pkg_or_name:
        return ""
    return f"https://www.qimai.cn/search/android/search/{urllib.parse.quote(pkg_or_name)}"


def _mark_incomplete(row):
    """检查结果行是否缺少关键字段，添加 incomplete 标记。
    完整结果需要：app_name, icon_url, category, download_url 全部非空，
    且 download_url 必须是稳定的商店页面（不能是 APK 直链）。
    对信息不完整 / 商店地址有问题的行，统一使用七麦搜索页作为兜底。"""
    # 先处理：download_url 若是 APK 直链，替换为七麦兜底
    # （APK 直链虽然非空，但对"商店地址"列来说是错的——过期即失效）
    dl = row.get("download_url", "")
    if dl and _is_apk_direct_url(dl):
        pkg = row.get("package_name") or row.get("app_name") or ""
        row["download_url"] = _qimai_fallback_url(pkg)
        # 标记来源，便于后续调试 / 记录
        orig_source = row.get("source", "")
        row["source"] = (orig_source + "+qimai_fallback").strip("+") if orig_source else "qimai_fallback"

    missing = []
    if not row.get("app_name") or row["app_name"] == "未找到":
        missing.append("app_name")
    if not row.get("icon_url"):
        missing.append("icon_url")
    if not row.get("category"):
        missing.append("category")
    if not row.get("download_url"):
        missing.append("download_url")

    # 任一字段缺失 → 兜底：download_url 确保指向七麦（即使本来是空的也填上）
    if missing:
        row["incomplete"] = True
        row["missing_fields"] = missing
        if not row.get("download_url"):
            pkg = row.get("package_name") or row.get("app_name") or ""
            if pkg:
                row["download_url"] = _qimai_fallback_url(pkg)
                # download_url 刚刚被填上，从 missing 里摘掉（其它字段仍缺失时 incomplete 保持 True）
                if "download_url" in row["missing_fields"]:
                    row["missing_fields"].remove("download_url")
    return row


def is_name_relevant(search_term, app_name):
    """判断App名称是否与搜索词相关。
    规则：
    - 完全相等 → 总是相关
    - 搜索词是 app 名的子串 → 相关（搜"斗地主"匹配"欢乐斗地主"）
    - app 名是搜索词的子串 → 看 app 名长度：中文≥2、英文≥4 才认为相关
      （避免"视频"/"QQ"这种通用短词误匹配任意含它的搜索词）
    """
    if not app_name or app_name == "未找到":
        return False
    s = search_term.lower().strip()
    a = app_name.lower().strip()
    if not s or not a:
        return False
    if s == a:
        return True
    if s in a:
        return True
    if a in s:
        # 是否含中文字符：中文短词信息量高（"微信"2字就足够），英文短词歧义大（"QQ"/"AI"）
        has_cjk = any('\u4e00' <= c <= '\u9fff' for c in a)
        min_len = 2 if has_cjk else 4
        return len(a) >= min_len
    return False


# ========== 按名称搜索 ==========

def search_apple_by_name(name, limit=3):
    """Apple iTunes Search API - 按名称搜索"""
    url = f"https://itunes.apple.com/search?term={urllib.parse.quote(name)}&country=cn&entity=software&limit={limit}"
    try:
        resp = _HTTP.get(url, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return []
        data = resp.json()
        results = []
        for app_info in data.get("results", []):
            bundle_id = app_info.get("bundleId", "")
            if not bundle_id:
                continue
            raw_name = app_info.get("trackName", "未知")
            app_name = clean_app_name(raw_name)
            raw_url = app_info.get("trackViewUrl", "")
            download_url = clean_ios_url(raw_url)
            icon_url = app_info.get("artworkUrl100", "") or app_info.get("artworkUrl60", "")
            genres = app_info.get("genres", [])
            category = genres[0] if genres else ""
            results.append({
                "package_name": bundle_id,
                "platform": "iOS",
                "app_name": app_name,
                "download_url": download_url,
                "icon_url": icon_url,
                "category": category,
            })
        return results
    except Exception:
        return []


def search_xiaomi_by_name(name):
    """小米应用商店 - 按名称搜索，返回 [{package_name, app_name}, ...]。

    统一契约：所有 *_by_name 函数返回 list[dict]，dict 至少含 package_name
    （app_name 可能为空，消费方需兼容）。

    2026 年小米网页侧已经做反爬：/search?keywords=... 会 302 到首页。
    如果不检测，页面里的 /details?id=... 链接其实是首页推荐，和搜索词无关，
    会把无关包名（如植物大战僵尸等）误当成候选，污染后面的匹配。
    """
    url = f"https://app.mi.com/search?keywords={urllib.parse.quote(name)}"
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return []
        # 被重定向到首页（或任何不带 keywords= 的 URL）→ 直接放弃，
        # 不把首页推荐的包名误当搜索结果。
        final_url = resp.url or ""
        if "keywords=" not in final_url or "/search" not in final_url:
            return []
        soup = BeautifulSoup(resp.text, _BS_PARSER)
        results = []
        seen = set()
        for link in soup.find_all("a", href=True):
            href = link.get("href", "")
            match = re.search(r'details\?id=([a-zA-Z0-9._\-]+)', href)
            if match:
                pkg = match.group(1)
                if pkg not in seen:
                    seen.add(pkg)
                    link_text = link.get_text(strip=True)
                    results.append({"package_name": pkg, "app_name": link_text or ""})
                    if len(results) >= 3:
                        break
        return results
    except Exception:
        return []


def search_tencent_by_name(name):
    """腾讯应用宝 - 按名称搜索，返回 [{package_name, app_name}, ...]。
    扫描整页 appdetail 链接及其文本，优先返回名称匹配的包名。
    应用宝页面体积大（约660KB），名称匹配的 app 可能出现在任何位置。
    """
    url = f"https://sj.qq.com/myapp/search.htm?kw={urllib.parse.quote(name)}"
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return []
        text = resp.text
        # 策略1：从 href="/appdetail/PKG">AppName</a> 中找名称匹配的包名
        # 应用宝页面中 appdetail 链接后紧跟着 app 名称作为链接文本
        pattern = r'appdetail/([a-zA-Z0-9._\-]+)"[^>]{0,100}>([^<]{1,40})</a>'
        results = []
        seen = set()
        for pkg, link_text in re.findall(pattern, text):
            if pkg in seen:
                continue
            link_text = link_text.strip()
            if link_text and is_name_relevant(name, link_text):
                seen.add(pkg)
                results.append({"package_name": pkg, "app_name": link_text})
                if len(results) >= 3:
                    return results
        # 不再无条件兜底：名称搜索没匹配时返回空，避免"光环助手"这类
        # 完全无关的 app 混入（与 iOS apple_results[:1] 已去除的反模式一致）
        return results
    except Exception:
        return []


def _wandoujia_headers():
    return {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "zh-CN,zh;q=0.9",
    }


def _parse_wandoujia_detail(detail_url, headers):
    """解析豌豆荚详情页，提取包名、App名、图标、分类"""
    try:
        dr = _HTTP.get(detail_url, headers=headers, timeout=HTTP_TIMEOUT)
        if dr.status_code != 200:
            return None
        ds = BeautifulSoup(dr.text, _BS_PARSER)

        # 包名：body data-pn 或 data-app-pname
        body = ds.find("body")
        pkg = body.get("data-pn", "") if body else ""
        if not pkg:
            el = ds.find(attrs={"data-app-pname": True})
            pkg = el.get("data-app-pname", "") if el else ""
        if not is_package_name(pkg):
            return None

        # App名称：标题格式 "微信下载2026安卓最新版_..."
        title_tag = ds.find("title")
        app_name = pkg
        if title_tag:
            t = title_tag.text.strip()
            m = re.match(r'^(.+?)下载', t)
            app_name = m.group(1).strip() if m else t.split("_")[0].strip()
            app_name = clean_app_name(app_name)

        # 图标：优先 25pp.com（豌豆荚App素材CDN）
        icon_url = ""
        for img in ds.find_all("img", src=True):
            if "25pp.com" in img["src"]:
                icon_url = img["src"]
                break

        # 分类：最细一级
        category = ""
        cat_links = ds.find_all("a", href=re.compile(r'/category/'))
        if cat_links:
            category = cat_links[-1].text.strip()

        return {
            "package_name": pkg,
            "platform": "Android",
            "app_name": app_name,
            "download_url": detail_url,
            "icon_url": icon_url,
            "category": category,
        }
    except Exception:
        return None


def search_wandoujia_by_name(name, limit=3):
    """豌豆荚两步搜索：名称→搜索页（含App名过滤）→详情页→包名+App信息"""
    headers = _wandoujia_headers()
    try:
        search_url = f"https://www.wandoujia.com/search?key={urllib.parse.quote(name)}"
        resp = _HTTP.get(search_url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return []
        soup = BeautifulSoup(resp.text, _BS_PARSER)

        # 提取 (app_id, app_name) 对，同一 ID 只取一次
        candidates = []  # [(app_id, link_text)]
        seen_ids = set()
        for a in soup.find_all("a", href=True):
            m = re.search(r'/apps/(\d+)', a["href"])
            if m:
                aid = m.group(1)
                link_text = a.get_text(strip=True)
                if aid not in seen_ids and link_text:
                    seen_ids.add(aid)
                    candidates.append((aid, link_text))

        if not candidates:
            return []

        # 优先取名称相关的结果
        relevant = [(aid, t) for aid, t in candidates if is_name_relevant(name, t)]
        if not relevant:
            relevant = candidates  # 无精确匹配时全取
        app_ids = [aid for aid, _ in relevant[:limit]]

        # 并发请求详情页
        detail_urls = [f"https://www.wandoujia.com/apps/{aid}" for aid in app_ids]
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(detail_urls)) as exe:
            for r in exe.map(lambda u: _parse_wandoujia_detail(u, headers), detail_urls):
                if r:
                    results.append(r)
        return results
    except Exception:
        return []


def search_wandoujia(package_name):
    """豌豆荚 - 按包名查询（搜索+精确匹配）"""
    headers = _wandoujia_headers()
    try:
        search_url = f"https://www.wandoujia.com/search?key={urllib.parse.quote(package_name)}"
        resp = _HTTP.get(search_url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, _BS_PARSER)

        app_ids = []
        seen = set()
        for a in soup.find_all("a", href=True):
            m = re.search(r'/apps/(\d+)', a["href"])
            if m:
                aid = m.group(1)
                if aid not in seen:
                    seen.add(aid)
                    app_ids.append(aid)
                    if len(app_ids) >= 5:
                        break

        for aid in app_ids:
            r = _parse_wandoujia_detail(f"https://www.wandoujia.com/apps/{aid}", headers)
            if r and r["package_name"] == package_name:
                r["source"] = "豌豆荚"
                return r
    except Exception:
        pass
    return None


def search_appchina_by_name(name, limit=3):
    """应用汇 - 按名称搜索，返回 [{package_name, app_name}, ...]。
    统一契约：所有 *_by_name 函数返回 list[dict]。"""
    url = f"http://www.appchina.com/search/?keywords={urllib.parse.quote(name)}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return []
        soup = BeautifulSoup(resp.text, _BS_PARSER)
        results = []
        seen = set()
        # 应用汇搜索结果页：每个app有个链接如 /app/com.xxx.xxx
        for link in soup.find_all("a", href=True):
            href = link.get("href", "")
            m = re.search(r'/app/([a-zA-Z0-9._\-]+)', href)
            if m:
                pkg = m.group(1)
                if is_package_name(pkg) and pkg not in seen:
                    seen.add(pkg)
                    link_text = link.get_text(strip=True)
                    results.append({"package_name": pkg, "app_name": link_text or ""})
                    if len(results) >= limit:
                        break
        return results
    except Exception:
        return []


def search_sogou_by_name(name, limit=3):
    """搜狗应用搜索 - 按名称搜索，返回完整结果（含 APK 下载地址和包名）
    URL: https://as.sogou.com/so?w=1459&pid=34&query={name}
    数据来源: SSR HTML，data-hd JSON 含下载重定向 URL
    包名从重定向 CDN URL 的 fsname 参数中提取（如 fsname=com.tencent.mm_8.0.70.apk）
    """
    import html as _html
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"}
    try:
        url = f"https://as.sogou.com/so?w=1459&pid=34&query={urllib.parse.quote(name)}"
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200 or len(resp.text) < 5000:
            return []
        # 每个 app 一个 data-hd JSON 块
        blocks = re.findall(r"data-hd='(\{[^']+\})'", resp.text)
        results = []
        for blk in blocks[:limit]:
            try:
                data = json.loads(_html.unescape(blk))
                app_name = clean_app_name(data.get("name", ""))
                if not app_name or not is_name_relevant(name, app_name):
                    continue
                icon_url = data.get("icon", "")
                down_redir = data.get("sogouHighdownUrl", "")
                if not down_redir:
                    continue
                # 跟进重定向取真实 CDN URL（fsname 参数含包名）
                r2 = _HTTP.get(down_redir, headers=headers, timeout=HTTP_TIMEOUT, allow_redirects=False)
                cdn_url = r2.headers.get("location", "")
                if not cdn_url:
                    continue
                fsname_m = re.search(r'fsname=([a-z][\w.]+?)_[\d.]+\.apk', cdn_url)
                if not fsname_m:
                    continue
                pkg_name = fsname_m.group(1)
                if not is_package_name(pkg_name):
                    continue
                results.append({
                    "package_name": pkg_name,
                    "platform":     "Android",
                    "app_name":     app_name,
                    "download_url": cdn_url,
                    "icon_url":     icon_url,
                    "category":     "",
                    "source":       "搜狗应用",
                })
            except Exception:
                continue
        return results
    except Exception:
        return []


def search_pp(package_name):
    """PP助手 (25pp.com) - 按包名查询"""
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        search_url = f"https://www.25pp.com/?q={urllib.parse.quote(package_name)}"
        resp = _HTTP.get(search_url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, _BS_PARSER)
        # 包名必须出现在页面中（data-app-pname 或 data-pn）
        if package_name not in resp.text:
            return None
        # 找到第一个包名匹配的详情链接
        detail_id = None
        app_name = ""
        icon_url = ""
        # 查找 data-app-pname 或 data-pn 属性匹配包名的元素
        for el in soup.find_all(attrs={"data-app-pname": package_name}):
            detail_href = el.get("href", "") or ""
            m = re.search(r'/detail/(\d+)', detail_href)
            if m:
                detail_id = m.group(1)
                break
        if not detail_id:
            for el in soup.find_all(attrs={"data-pn": package_name}):
                detail_href = el.get("href", "") or ""
                m = re.search(r'/detail/(\d+)', detail_href)
                if m:
                    detail_id = m.group(1)
                    break
        # 如果没有精确属性，从链接中找
        if not detail_id:
            for a in soup.find_all("a", href=True):
                if re.search(r'/detail/\d+', a.get("href", "")):
                    m = re.search(r'/detail/(\d+)', a["href"])
                    if m:
                        detail_id = m.group(1)
                        break
        if not detail_id:
            return None
        detail_url = f"https://www.25pp.com/detail/{detail_id}"
        # 获取详情页取 App 名和图标
        dr = _HTTP.get(detail_url, headers=headers, timeout=HTTP_TIMEOUT)
        if dr.status_code == 200:
            ds = BeautifulSoup(dr.text, _BS_PARSER)
            # App名: <span class="title" itemprop="name">
            name_tag = ds.find("span", class_="title") or ds.find("p", class_="title")
            if name_tag:
                app_name = clean_app_name(name_tag.get_text(strip=True))
            # 图标: <img> in <div class="app-icon">
            icon_div = ds.find("div", class_="app-icon")
            if icon_div:
                img = icon_div.find("img")
                if img:
                    icon_url = img.get("src", "")
        if not app_name:
            return None
        return {
            "source": "PP助手",
            "app_name": app_name,
            "download_url": detail_url,
            "icon_url": icon_url,
            "category": "",
        }
    except Exception:
        pass
    return None


def search_flyme(package_name):
    """魅族 Flyme 应用商店 - 按包名查询（SSR HTML，无需认证）"""
    url = f"https://app.flyme.cn/apps/public/detail?package_name={urllib.parse.quote(package_name)}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200 or len(resp.text) < 5000:
            return None
        # App 名和分类藏在 <input id="count"> 的 data-* 属性里
        m = re.search(
            r'<input[^>]+id="count"[^>]+data-cname="([^"]*)"[^>]+data-cid="\d+"[^>]+data-name="([^"]+)"',
            resp.text
        )
        if not m or not m.group(2):
            return None
        import html as _html
        app_name = clean_app_name(_html.unescape(m.group(2)))
        category = _html.unescape(m.group(1))
        if not app_name:
            return None
        # 图标
        icon_m = re.search(r'<img\s+class="app_img"\s+src="([^"]+)"', resp.text)
        icon_url = icon_m.group(1) if icon_m else ""
        return {
            "source":       "魅族Flyme",
            "app_name":     app_name,
            "download_url": f"https://app.flyme.cn/apps/public/detail?package_name={urllib.parse.quote(package_name)}",
            "icon_url":     icon_url,
            "category":     category,
        }
    except Exception:
        pass
    return None


def search_appchina(package_name):
    """应用汇 - 按包名查询"""
    url = f"http://www.appchina.com/app/{package_name}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, _BS_PARSER)
        name_tag = soup.find("h1", class_="app-name")
        if not name_tag:
            return None
        app_name = name_tag.get_text(strip=True)
        if not app_name or "404" in app_name:
            return None
        icon_tag = soup.find("img", class_="Content_Icon")
        icon_url = icon_tag.get("src", "") if icon_tag else ""
        m = re.search(r'McDonald[/\w]*?/(\d{6,8})/', resp.text)
        apk_direct_url = (
            f"http://mobile.d.appchina.com/McDonald/r/{m.group(1)}/{package_name}.vapk"
            if m else ""
        )
        return {
            "source": "应用汇",
            "app_name": app_name,
            "download_url": url,
            "icon_url": icon_url,
            "category": "",
            "apk_direct_url": apk_direct_url,
        }
    except Exception:
        pass
    return None


def _get_wandoujia_apk_url(package_name):
    """从豌豆荚获取 APK 直链（仅用于多来源模式）"""
    wdj_headers = _wandoujia_headers()
    req_headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        search_url = f"https://www.wandoujia.com/search?key={urllib.parse.quote(package_name)}"
        resp = _HTTP.get(search_url, headers=wdj_headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return ""
        soup = BeautifulSoup(resp.text, _BS_PARSER)
        for a in soup.find_all("a", href=True):
            m = re.search(r'/apps/(\d+)', a["href"])
            if m:
                wdj_id = m.group(1)
                r = _parse_wandoujia_detail(
                    f"https://www.wandoujia.com/apps/{wdj_id}", wdj_headers
                )
                if r and r.get("package_name") == package_name:
                    dl_url = (
                        f"https://www.wandoujia.com/apps/{wdj_id}/download/dot"
                        f"?ch=detail_qr_dl&pos=detail-ndownload-{package_name}"
                    )
                    dl_r = _HTTP.get(
                        dl_url, headers=req_headers,
                        allow_redirects=True, timeout=HTTP_TIMEOUT * 2
                    )
                    if ".apk" in dl_r.url:
                        return dl_r.url
                    break
    except Exception:
        pass
    return ""


def _fingerprint_str(cert, hash_algo):
    return ":".join(f"{b:02X}" for b in cert.fingerprint(hash_algo))


def _extract_cert_fingerprints(apk_url):
    """通过 Range 请求只下载 APK 证书部分，同时计算 SHA1 和 SHA256 指纹。
    优先尝试 APK v2/v3 签名块（现代 APK），回退到 v1 META-INF/*.RSA。
    返回 (sha1_str, sha256_str) 元组，失败时返回 ("", "")。"""
    if not HAS_CRYPTOGRAPHY:
        return "", ""
    from cryptography import x509 as crypto_x509
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        # 1. 获取文件大小
        r = _HTTP.head(apk_url, headers=headers, timeout=10, allow_redirects=True)
        fsize = int(r.headers.get("Content-Length", 0))
        if not fsize:
            return "", ""

        # 2. 下载末尾 64KB，找 ZIP EOCD 和 Central Directory 位置
        tail_size = min(65536, fsize)
        r = _HTTP.get(
            apk_url,
            headers={**headers, "Range": f"bytes={fsize - tail_size}-{fsize - 1}"},
            timeout=15, allow_redirects=True
        )
        tail = r.content
        eocd_pos = tail.rfind(b'PK\x05\x06')
        if eocd_pos < 0:
            return "", ""
        eocd = tail[eocd_pos:]
        cd_size   = struct.unpack_from('<L', eocd, 12)[0]
        cd_offset = struct.unpack_from('<L', eocd, 16)[0]

        # ── 尝试 APK v2/v3 签名块（位于 Central Directory 之前）─────────
        APK_SIG_MAGIC = b'APK Sig Block 42'
        APK_V2_ID = 0x7109871a
        APK_V3_ID = 0xf05368c0
        if cd_offset > 32:
            r2 = _HTTP.get(
                apk_url,
                headers={**headers, "Range": f"bytes={cd_offset - 32}-{cd_offset - 1}"},
                timeout=10, allow_redirects=True
            )
            footer = r2.content
            if len(footer) >= 24 and footer[-16:] == APK_SIG_MAGIC:
                block_size = struct.unpack_from('<Q', footer, len(footer) - 24)[0]
                block_start = cd_offset - block_size - 8
                if block_start >= 0:
                    r3 = _HTTP.get(
                        apk_url,
                        headers={**headers, "Range": f"bytes={block_start}-{cd_offset - 1}"},
                        timeout=20, allow_redirects=True
                    )
                    block = r3.content
                    bp = 8
                    end_pairs = len(block) - 24
                    while bp < end_pairs:
                        pl = struct.unpack_from('<Q', block, bp)[0]
                        pid = struct.unpack_from('<I', block, bp + 8)[0]
                        if pid in (APK_V2_ID, APK_V3_ID):
                            val = block[bp + 12: bp + 8 + pl]
                            try:
                                sp = 4
                                slen = struct.unpack_from('<I', val, sp)[0]
                                sd = val[sp + 4: sp + 4 + slen]
                                sdlen = struct.unpack_from('<I', sd, 0)[0]
                                signed = sd[4: 4 + sdlen]
                                dlen = struct.unpack_from('<I', signed, 0)[0]
                                cs = 4 + dlen
                                cp = cs + 4
                                clen = struct.unpack_from('<I', signed, cp)[0]
                                cert_der = signed[cp + 4: cp + 4 + clen]
                                cert = crypto_x509.load_der_x509_certificate(cert_der)
                                return (
                                    _fingerprint_str(cert, crypto_hashes.SHA1()),
                                    _fingerprint_str(cert, crypto_hashes.SHA256()),
                                )
                            except Exception:
                                pass
                        bp += 8 + pl

        # ── 回退到 v1 签名：META-INF/*.RSA / *.DSA / *.EC ────────────────
        r = _HTTP.get(
            apk_url,
            headers={**headers, "Range": f"bytes={cd_offset}-{cd_offset + cd_size - 1}"},
            timeout=15, allow_redirects=True
        )
        cd = r.content

        pos = 0
        cert_info = None
        while pos < len(cd) - 4:
            if cd[pos:pos + 4] != b'PK\x01\x02':
                break
            fn_len  = struct.unpack_from('<H', cd, pos + 28)[0]
            ex_len  = struct.unpack_from('<H', cd, pos + 30)[0]
            cm_len  = struct.unpack_from('<H', cd, pos + 32)[0]
            lh_off  = struct.unpack_from('<L', cd, pos + 42)[0]
            comp_sz = struct.unpack_from('<L', cd, pos + 20)[0]
            method  = struct.unpack_from('<H', cd, pos + 10)[0]
            fname   = cd[pos + 46:pos + 46 + fn_len].decode('utf-8', errors='replace')
            if re.match(r'META-INF/[^/]+\.(RSA|DSA|EC)$', fname, re.I):
                cert_info = (lh_off, comp_sz, method)
                break
            pos += 46 + fn_len + ex_len + cm_len

        if not cert_info:
            return "", ""
        lh_off, comp_sz, method = cert_info

        end = lh_off + 30 + 512 + comp_sz
        r = _HTTP.get(
            apk_url,
            headers={**headers, "Range": f"bytes={lh_off}-{end}"},
            timeout=15, allow_redirects=True
        )
        local = r.content
        fn_l = struct.unpack_from('<H', local, 26)[0]
        ex_l = struct.unpack_from('<H', local, 28)[0]
        data = local[30 + fn_l + ex_l: 30 + fn_l + ex_l + comp_sz]
        if method == 8:
            data = zlib.decompress(data, -15)

        certs = crypto_pkcs7.load_der_pkcs7_certificates(data)
        if not certs:
            return "", ""
        return (
            _fingerprint_str(certs[0], crypto_hashes.SHA1()),
            _fingerprint_str(certs[0], crypto_hashes.SHA256()),
        )
    except Exception:
        return "", ""


def _has_android_match(ios_name, android_results):
    """判断 iOS app 是否已有对应的安卓版本（严格匹配，防止"抖音"误匹配"抖音极速版"）"""
    for ar in android_results:
        an = ar["app_name"].strip()
        ios = ios_name.strip()
        if ios == an:
            return True
        # 长短比例超过 60% 才算相关（避免子集误判）
        shorter, longer = sorted([ios, an], key=len)
        if shorter and shorter in longer and len(shorter) >= len(longer) * 0.6:
            return True
    return False


def query_by_name(name, android_store_order=None, exact_search=False):
    """按App名称搜索，iOS + Android 互补：
    1. 并发：Apple API（iOS）+ Android 各商店（多商店回退）
    2. 对没有安卓对应的 iOS app，补搜安卓商店
    3. 对安卓结果，用优先商店丰富下载链接/分类/图标
    exact_search=True 时只返回与搜索词完全一致的 App
    """
    if android_store_order is None:
        android_store_order = get_ranked_store_order()

    # === 第一轮：并发 iOS + Android 多商店搜索 ===
    # 按 android_store_order 的顺序依次尝试，直到找到结果为止
    def _search_android_by_name(name, store_order):
        """按商店优先级搜索，所有商店并发查，按优先级取结果。
        策略1: 各中国商店名称搜索（全部并发）
        策略2: Bing 搜索 → 获取包名 → 查商店详情
        """
        # 策略1：全部商店并发名称搜索（不再逐个串行）
        name_funcs = {}
        for sid in store_order:
            func = STORE_NAME_SEARCH_FUNCS.get(sid)
            if func:
                name_funcs[sid] = func

        if name_funcs:
            with concurrent.futures.ThreadPoolExecutor(
                    max_workers=len(name_funcs)) as exe:
                futures = {sid: exe.submit(fn, name)
                           for sid, fn in name_funcs.items()}
                # 收集所有商店的相关结果，优先返回精确匹配（防止"豆包爱学"遮蔽"豆包"）
                # 精确匹配：app_name 与搜索词完全相同（忽略大小写/前后空格）
                exact_results = []
                partial_results = []
                for sid in store_order:
                    if sid not in futures:
                        continue
                    try:
                        results = futures[sid].result(timeout=HTTP_TIMEOUT + 2)
                        if not results:
                            continue
                        for r in results:
                            an = r.get("app_name", "").strip().lower()
                            nm = name.strip().lower()
                            if an == nm:
                                exact_results.append(r)
                            elif is_name_relevant(name, r.get("app_name", "")):
                                partial_results.append(r)
                    except Exception:
                        continue
                if exact_results:
                    return exact_results
                if partial_results:
                    return partial_results

        # 策略2：Bing 搜索包名（国内可用）
        bing_pkgs = search_bing_for_android_package(name, 5)

        for pkg_list in [bing_pkgs]:
            for pkg in (pkg_list or []):
                # 并发查各商店
                with concurrent.futures.ThreadPoolExecutor(
                        max_workers=len(store_order)) as exe2:
                    pkg_futures = {}
                    for sid in store_order:
                        func = STORE_SEARCH_FUNCS.get(sid)
                        if func:
                            pkg_futures[sid] = exe2.submit(func, pkg)
                    for sid in store_order:
                        if sid not in pkg_futures:
                            continue
                        try:
                            r = pkg_futures[sid].result(timeout=HTTP_TIMEOUT + 1)
                            if r and is_name_relevant(name, r.get("app_name", "")):
                                return [{
                                    "package_name": pkg,
                                    "platform": "Android",
                                    "app_name": r["app_name"],
                                    "download_url": r["download_url"],
                                    "icon_url": r.get("icon_url", ""),
                                    "category": r.get("category", ""),
                                }]
                        except Exception:
                            continue
        return []

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        future_ios = executor.submit(search_apple_by_name, name, 5)
        future_android = executor.submit(_search_android_by_name, name, android_store_order)
        apple_results = future_ios.result()
        android_results_raw = future_android.result()

    # iOS 相关结果（最多3个）
    if exact_search:
        relevant_ios = [r for r in apple_results if r["app_name"].strip().lower() == name.strip().lower()]
    else:
        relevant_ios = [r for r in apple_results if is_name_relevant(name, r["app_name"])]
    # 不再无条件回退取第一条：当没有任何相关结果时，宁可不返回也不返回无关 app
    # （之前的回退会导致搜"斗地主"却返回"腾讯视频"等无关结果）
    relevant_ios = relevant_ios[:3]

    # === 第二轮：为没有安卓对应的 iOS app 补搜安卓商店 ===
    # 策略1: iOS bundle ID 直查安卓商店（最可靠，跨平台包名往往相同）
    # 策略2: 回退到名称搜索
    ios_without_android = [
        r for r in relevant_ios
        if not _has_android_match(r["app_name"], android_results_raw)
    ]

    def _try_android_for_ios(ios_r):
        """为 iOS app 找对应安卓信息（并发查商店）"""
        bid = ios_r["package_name"]
        ios_name = ios_r["app_name"]

        # 策略1: iOS bundle ID 并发查所有安卓商店
        store_funcs = [(sid, STORE_SEARCH_FUNCS[sid])
                       for sid in android_store_order
                       if sid in STORE_SEARCH_FUNCS]
        if store_funcs:
            with concurrent.futures.ThreadPoolExecutor(
                    max_workers=len(store_funcs)) as exe:
                futs = {sid: exe.submit(fn, bid) for sid, fn in store_funcs}
                for sid in android_store_order:
                    if sid not in futs:
                        continue
                    try:
                        r = futs[sid].result(timeout=HTTP_TIMEOUT + 1)
                        if r and is_name_relevant(ios_name, r.get("app_name", "")):
                            return {
                                "package_name": bid,
                                "platform": "Android",
                                "app_name": r["app_name"],
                                "download_url": r["download_url"],
                                "icon_url": r.get("icon_url", ""),
                                "category": r.get("category", ""),
                            }
                    except Exception:
                        continue

        # 策略2: 名称搜索（内部已并发化）
        for ar in (_search_android_by_name(ios_name, android_store_order) or []):
            if _has_android_match(ios_name, [ar]):
                return ar

        # 策略3: 用包名走完整 query_single 链（含搜索引擎反查 + qimai兜底）
        # 注意：若结果仍是 qimai_hint 且包名与 iOS bundle 相同，说明没找到真正的 Android 包，丢弃
        rows = query_single(bid, android_store_order)
        for r in (rows or []):
            if r.get("platform") == "Android":
                # 过滤掉"用 iOS bundle ID 伪装成 Android 包"的 qimai_hint
                if r.get("source") == "qimai_hint" and r.get("package_name") == bid:
                    continue
                return r

        return None

    extra_android = []
    if ios_without_android:
        with concurrent.futures.ThreadPoolExecutor(
                max_workers=min(3, len(ios_without_android))) as exe:
            for result in exe.map(_try_android_for_ios, ios_without_android[:3]):
                if result:
                    extra_android.append(result)

    all_android = android_results_raw + extra_android

    # === 第三轮：按用户商店优先级丰富安卓结果（下载链接、分类、图标）===
    # 规则：遍历 android_store_order，遇到 wandoujia 时说明豌豆荚就是优先商店，
    # 保留已有的豌豆荚数据；遇到其他商店时尝试获取并替换。
    def enrich_android(ar):
        """按用户配置优先级替换/补全安卓信息，并跨商店取最简洁的 app 名称"""
        pkg = ar["package_name"]
        all_names = [ar.get("app_name", "")]
        enriched = dict(ar)
        found_primary = False

        for store_id in android_store_order:
            if store_id == "wandoujia":
                all_names.append(ar.get("app_name", ""))
                if not found_primary:
                    found_primary = True
                continue
            func = STORE_SEARCH_FUNCS.get(store_id)
            if not func:
                continue
            try:
                r = func(pkg)
                if r:
                    all_names.append(r.get("app_name", ""))
                    if not found_primary:
                        found_primary = True
                        enriched["download_url"] = r["download_url"]
                        if r.get("category"):
                            enriched["category"] = r["category"]
                        if r.get("icon_url"):
                            enriched["icon_url"] = r["icon_url"]
                    else:
                        # 补全缺失字段
                        if r.get("category") and not enriched.get("category"):
                            enriched["category"] = r["category"]
                        if r.get("icon_url") and not enriched.get("icon_url"):
                            enriched["icon_url"] = r["icon_url"]
            except Exception:
                continue

        # 第一个名字来自搜索命中（已通过 is_name_relevant 检验），作为基准名
        best = pick_best_name(all_names, primary=ar.get("app_name"))
        if best:
            enriched["app_name"] = best
        return enriched

    # 精确搜索时：安卓结果也过滤
    # 允许 "豆包app" 匹配 "豆包"（商店会在名称后加 "app" 后缀）
    if exact_search:
        n_lower = name.strip().lower()
        def _android_exact_match(app_name):
            a = app_name.strip().lower()
            if a == n_lower:
                return True
            # 允许 "豆包app" "豆包App" "豆包 - AI助手" 等常见后缀变体
            if a.startswith(n_lower) and (len(a) == len(n_lower) or a[len(n_lower)] in (' ', '-', '（', '(', 'a')):
                return True
            return False
        all_android = [r for r in all_android if _android_exact_match(r["app_name"])]

    enriched_map = {}
    if all_android:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(4, len(all_android))) as exe:
            for ar, enriched in zip(all_android, exe.map(enrich_android, all_android)):
                enriched_map[ar["package_name"]] = enriched

    def get_android(ar):
        return enriched_map.get(ar["package_name"], ar)

    # === 组装结果：iOS + 配对安卓，成对排列 ===
    # 用 (package_name, platform) 做唯一键，防止 iOS/Android 包名相同时互相覆盖
    rows = []
    seen_keys = set()   # (pkg, platform)
    used_android = set()  # 已配对的安卓包名

    for ios_r in relevant_ios:
        ios_key = (ios_r["package_name"], "iOS")
        if ios_key in seen_keys:
            continue
        seen_keys.add(ios_key)
        rows.append(_mark_incomplete({
            "package_name": ios_r["package_name"],
            "platform": "iOS",
            "app_name": ios_r["app_name"],
            "download_url": ios_r["download_url"],
            "icon_url": ios_r.get("icon_url", ""),
            "category": ios_r.get("category", ""),
        }))
        # 找配对安卓（严格名称匹配）
        for ar in all_android:
            a_key = (ar["package_name"], "Android")
            if a_key not in seen_keys and _has_android_match(ios_r["app_name"], [ar]):
                seen_keys.add(a_key)
                used_android.add(ar["package_name"])
                rows.append(_mark_incomplete(get_android(ar)))
                break

    # 未配对的安卓单独追加
    for ar in all_android:
        a_key = (ar["package_name"], "Android")
        if a_key not in seen_keys:
            seen_keys.add(a_key)
            rows.append(_mark_incomplete(get_android(ar)))

    if not rows:
        return [{
            "package_name": name,
            "platform": "未知",
            "app_name": "未找到",
            "download_url": "",
            "icon_url": "",
            "category": "",
        }]
    return rows


# ========== 查询逻辑 ==========

HTTP_TIMEOUT = 5 if os.name == "nt" else 3  # 统一超时时间（秒）：Windows 的 DNS/TLS/防火墙比 Mac 慢，放宽到 5s 以保住小米/应用宝


def search_xiaomi(package_name):
    """小米应用商店。

    2026 年起 app.mi.com/details?id=... 会对未上架/被下架/某些包名直接 302 到首页。
    落到首页的 HTML 里 title 是"手机游戏应用商店_软件商店app下载-小米应用商店"，
    不做重定向检测的话需要依赖 skip_keywords 兜底。这里直接检查 final_url，
    不再包含 details?id 就跳过。
    """
    url = f"https://app.mi.com/details?id={package_name}"
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return None
        # 被重定向到首页 → 这个包名在小米拿不到有效结果
        final_url = resp.url or ""
        if "details?id=" not in final_url and "/details" not in final_url:
            return None
        soup = BeautifulSoup(resp.text, _BS_PARSER)
        title_tag = soup.find("title")
        if not title_tag:
            return None
        title_text = title_tag.text
        if "404" in title_text or "错误" in title_text:
            return None
        app_name = title_text.split("-")[0].strip()
        skip_keywords = ["应用商店", "软件商店", "游戏商店", "手机游戏"]
        if app_name and not any(kw in app_name for kw in skip_keywords):
            icon_url = ""
            icon_tag = soup.find("img", class_="yellow-flower") or soup.find("img", attrs={"alt": app_name})
            if icon_tag and icon_tag.get("src"):
                icon_url = icon_tag["src"]
                if icon_url.startswith("//"):
                    icon_url = "https:" + icon_url

            # 提取分类（面包屑导航：首页 > 分类 > App名）
            category = ""
            breadcrumbs = soup.find_all("a", href=re.compile(r'/category/'))
            if breadcrumbs:
                category = breadcrumbs[0].get_text(strip=True)

            return {
                "source": "小米应用商店",
                "app_name": app_name,
                "download_url": url,
                "icon_url": icon_url,
                "category": category,
            }
    except Exception:
        pass
    return None


def search_tencent(package_name):
    """腾讯应用宝 - 优先从 __NEXT_DATA__ JSON 提取字段，HTML 回退。

    2026 版 sj.qq.com 是 Next.js SSR 页，<a href="/category/..."> 已经没了，
    分类（cate_name/cate_name_new）、名称、图标等都在 __NEXT_DATA__ 里。
    """
    url = f"https://sj.qq.com/appdetail/{package_name}"
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return None

        app_name = ""
        icon_url = ""
        category = ""

        # ---- 首选：__NEXT_DATA__ JSON ----
        try:
            m = re.search(
                r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>',
                resp.text, re.S,
            )
            if m:
                nd = json.loads(m.group(1))
                comps = (nd.get("props", {})
                           .get("pageProps", {})
                           .get("dynamicCardResponse", {})
                           .get("data", {})
                           .get("components", []))
                # 第一个包含 itemData[0].pkg_name == package_name 的 component 就是 app 详情
                for c in comps:
                    items = (c.get("data") or {}).get("itemData") or []
                    if not items:
                        continue
                    first = items[0] if isinstance(items[0], dict) else {}
                    if first.get("pkg_name") == package_name:
                        app_name = first.get("name", "") or app_name
                        icon_url = first.get("icon", "") or icon_url
                        category = (first.get("cate_name_new")
                                    or first.get("cate_name")
                                    or "")
                        break
        except Exception:
            pass

        # ---- 回退：HTML 解析 ----
        if not app_name or not icon_url:
            soup = BeautifulSoup(resp.text, _BS_PARSER)
            if not app_name:
                title_tag = soup.find("title")
                if title_tag:
                    title_text = title_tag.text.strip()
                    app_name = title_text.split("-")[0].strip()
                    for suffix in ["下载安装", "下载", "安装", "app"]:
                        if app_name.lower().endswith(suffix):
                            app_name = app_name[:-len(suffix)].strip()
                            break
            if not icon_url:
                icon_tag = soup.find("img", class_="det-icon") or soup.find("img", attrs={"alt": app_name})
                if icon_tag and icon_tag.get("src"):
                    icon_url = icon_tag["src"]
                    if icon_url.startswith("//"):
                        icon_url = "https:" + icon_url

        if app_name and app_name != "应用宝" and "404" not in app_name:
            # icon 如果只是 http:// 协议，升级到 https 方便前端展示
            if icon_url.startswith("http://"):
                icon_url = "https://" + icon_url[len("http://"):]
            return {
                "source": "腾讯应用宝",
                "app_name": app_name,
                "download_url": url,
                "icon_url": icon_url,
                "category": category,
            }
    except Exception:
        pass
    return None


def search_apple(bundle_id):
    """Apple App Store (iTunes API) - 按 bundle ID 查询"""
    url = f"https://itunes.apple.com/lookup?bundleId={bundle_id}&country=cn"
    try:
        resp = _HTTP.get(url, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return None
        data = resp.json()
        if data.get("resultCount", 0) > 0:
            app_info = data["results"][0]
            raw_name = app_info.get("trackName", "未知")
            app_name = clean_app_name(raw_name)
            raw_url = app_info.get("trackViewUrl", "")
            download_url = clean_ios_url(raw_url)
            icon_url = app_info.get("artworkUrl100", "") or app_info.get("artworkUrl60", "")
            # iTunes API 直接提供分类
            genres = app_info.get("genres", [])
            category = genres[0] if genres else ""
            return {
                "source": "Apple App Store",
                "app_name": app_name,
                "download_url": download_url,
                "icon_url": icon_url,
                "category": category,
            }
    except Exception:
        pass
    return None


def search_apple_by_numid(app_id):
    """Apple App Store (iTunes API) - 按数字 App ID 查询（如 414478124）"""
    url = f"https://itunes.apple.com/lookup?id={app_id}&country=cn"
    try:
        resp = _HTTP.get(url, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            return None
        data = resp.json()
        if data.get("resultCount", 0) > 0:
            app_info = data["results"][0]
            raw_name = app_info.get("trackName", "未知")
            app_name = clean_app_name(raw_name)
            raw_url = app_info.get("trackViewUrl", "")
            download_url = clean_ios_url(raw_url)
            icon_url = app_info.get("artworkUrl100", "") or app_info.get("artworkUrl60", "")
            genres = app_info.get("genres", [])
            category = genres[0] if genres else ""
            bundle_id = app_info.get("bundleId", app_id)
            return {
                "package_name": bundle_id,
                "platform": "iOS",
                "app_name": app_name,
                "download_url": download_url,
                "icon_url": icon_url,
                "category": category,
            }
    except Exception:
        pass
    return None


# 默认安卓商店优先级（初始顺序，基于经验：小米和腾讯命中率最高）
DEFAULT_ANDROID_STORES = ["xiaomi", "tencent", "wandoujia", "appchina", "pp", "sogou"]

# ── 商店自动排名统计 ──────────────────────────────────────────────────
# 追踪每个商店的命中率和平均响应时间，自动调整查询优先级
_store_stats_lock = threading.Lock()
_store_stats = {}  # store_id → {"hits": int, "misses": int, "total_time": float, "queries": int}

def _record_store_stat(store_id, hit, elapsed):
    """记录一次商店查询的结果（线程安全）"""
    with _store_stats_lock:
        if store_id not in _store_stats:
            _store_stats[store_id] = {"hits": 0, "misses": 0, "total_time": 0.0, "queries": 0}
        s = _store_stats[store_id]
        s["queries"] += 1
        s["total_time"] += elapsed
        if hit:
            s["hits"] += 1
        else:
            s["misses"] += 1

def get_ranked_store_order():
    """根据统计数据自动排序商店：命中率高+速度快的排前面。
    统计不足时使用默认顺序。"""
    with _store_stats_lock:
        if not _store_stats:
            return list(DEFAULT_ANDROID_STORES)
        # 至少需要每个商店 5 次查询才开始排名
        min_queries = min((s.get("queries", 0) for s in _store_stats.values()), default=0)
        if min_queries < 5:
            return list(DEFAULT_ANDROID_STORES)
        ranked = []
        for sid in DEFAULT_ANDROID_STORES:
            s = _store_stats.get(sid)
            if s and s["queries"] > 0:
                hit_rate = s["hits"] / s["queries"]
                avg_time = s["total_time"] / s["queries"]
                # 排序分数：命中率权重 0.7，速度权重 0.3（归一化到0-1，3秒为最慢基准）
                speed_score = max(0, 1.0 - avg_time / 3.0)
                score = hit_rate * 0.7 + speed_score * 0.3
                ranked.append((sid, score))
            else:
                ranked.append((sid, 0.5))  # 无数据的商店排中间
        ranked.sort(key=lambda x: -x[1])
        return [sid for sid, _ in ranked]

STORE_SEARCH_FUNCS = {
    "xiaomi":    search_xiaomi,
    "tencent":   search_tencent,
    "wandoujia": search_wandoujia,
    "appchina":  search_appchina,
    "pp":        search_pp,
    "flyme":     search_flyme,
}

STORE_NAMES = {
    "xiaomi":    "小米应用商店",
    "tencent":   "腾讯应用宝",
    "wandoujia": "豌豆荚",
    "appchina":  "应用汇",
    "pp":        "PP助手",
    "flyme":     "魅族Flyme",
    "sogou":     "搜狗应用",
}

# 名称搜索函数：用于 query_by_name 的多商店回退
# 注意：小米/腾讯/应用汇的搜索页均为 JS 渲染，静态 HTML 只有首页推荐内容，无法用于名称搜索
# 豌豆荚和搜狗支持服务端渲染搜索结果，能正常提取包名
STORE_NAME_SEARCH_FUNCS = {
    "wandoujia": search_wandoujia_by_name,
    "sogou":     lambda name: search_sogou_by_name(name, 3),
}


# 搜索引擎反查结果内存缓存（包名 → (app_name, ref_url)），避免重复请求导致限流
_search_engine_cache = {}
_search_engine_cache_lock = threading.Lock()  # 多线程并发查询时保护缓存读写
# 搜狗/360搜索上次请求时间戳，用于限速（最小间隔3秒，避免IP封锁）
_sogou_search_last_call = [0.0]
_360_search_last_call = [0.0]
# 搜狗持久会话（保持cookie，减少触发反爬）
_sogou_session = None
_sogou_session_lock = threading.Lock()  # 避免并发首次初始化时重复访问首页

# APK/应用分发站域名关键词（优先选这些作为参考链接）
_PREFERRED_APK_DOMAINS = ("apk", "soft", "shouji", "game", "app", "down")

# 投票提取 app 名称时用于过滤"描述性短语"的正则（一款/一个/一种 + 常见动词/形容词）
# 在 _extract_app_name_from_titles 的热点循环中使用，提到模块级避免每次 re.compile
_DESC_PHRASE_RE = re.compile(
    r'^一[款个种类]'               # 一款、一个、一种、一类
    r'|[可使能介功特]'             # 可以、使用、能够、介绍、功能、特点
    r'|(?:查询|查看|管理|帮助|支持|提供|实现|完成|解决|适合|满足)'
)

# iOS 数字 ID 识别（批量查询解析时每条输入都要匹配，提到模块级避免重复编译）
_IOS_ID_RE = re.compile(r'^\d{6,12}$')


def _get_sogou_session():
    """获取或初始化搜狗持久会话（首次访问首页以建立正常cookie）。
    使用 double-check locking：大多数情况下已初始化，无锁快速返回；
    只有首次竞争时才进入锁，避免多个线程同时访问首页被反爬识别为异常。
    """
    global _sogou_session
    if _sogou_session is not None:
        return _sogou_session
    with _sogou_session_lock:
        if _sogou_session is not None:
            return _sogou_session
        s = requests.Session()
        s.headers.update({
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://www.sogou.com/",
        })
        try:
            s.get("https://www.sogou.com/", timeout=HTTP_TIMEOUT)
            time.sleep(1)
        except Exception:
            pass
        _sogou_session = s
        return s


def _extract_app_name_from_titles(titles, snippets, package_name):
    """从标题和摘要列表中投票提取最可信的应用名称。返回 app_name 或 None。"""
    from collections import Counter
    all_texts = titles + snippets
    full_text = " ".join(all_texts)
    candidates = []

    # 模式0: 标题/摘要中"应用名(package_name)"格式（APK站最常见，权重5）
    # 例：com.tencent.mm apk下载-微信(com.tencent.mm)下载大全 → 提取"微信"
    # 只匹配括号前紧邻的纯中文词（2-8字），避免匹配带连字符的噪音
    pkg_escaped = re.escape(package_name)
    for text in all_texts:
        m = re.search(r'([\u4e00-\u9fff]{2,8})\(' + pkg_escaped + r'\)', text)
        if m:
            candidates.append((m.group(1).strip(), 5))

    # 模式1a: 摘要中"软件中文名为"X"" / "名称为"X""（明确说明，权重5）
    for text in snippets:
        for pat in [r'软件中文名为["""]([\u4e00-\u9fffA-Za-z0-9\-_]{2,15})["""]',
                    r'中文名为["""]([\u4e00-\u9fffA-Za-z0-9\-_]{2,15})["""]',
                    r'名称[：:为]["""]([\u4e00-\u9fffA-Za-z0-9\-_]{2,15})["""]']:
            for m in re.finditer(pat, text):
                name_cand = m.group(1).strip()
                if 2 <= len(name_cand) <= 15:
                    candidates.append((name_cand, 5))

    # 模式1b: "是X的包名" / "是X软件/应用的包名"（权重4）
    for m in re.finditer(
        r'是\s*([^\s，。,（(]{2,12}?)\s*(?:软件|应用|App|APP)?的?包名', full_text
    ):
        name_cand = m.group(1).strip()
        if 2 <= len(name_cand) <= 12:
            candidates.append((name_cand, 4))

    # 模式2: 标题首段中文词 + 后跟分隔符（APK站/应用市场常见格式，权重2）
    # 过滤：候选词必须≤8个汉字，避免提取"谷歌微信最新版本下载安装"这类噪声
    for text in titles:
        m = re.match(r'^([\u4e00-\u9fff]{2,8})\s*[-–|/（(]', text)
        if m:
            name_cand = m.group(1).strip()
            if 2 <= len(name_cand) <= 8:
                candidates.append((name_cand, 2))

    # 模式3: 标题中"X APK下载/最新版/官方版"（权重2，同样限制汉字长度）
    for text in titles:
        # 匹配"应用名[免费/专业/最新/官方]下载/APK"，中间修饰词不算入名称
        m = re.match(r'^([\u4e00-\u9fff]{2,8}?)\s*(?:免费|专业版?|正版?|精品|官方版?|最新)?\s*(?:APK|apk|下载|最新版|官方|正式版)', text)
        if m:
            name_cand = m.group(1).strip()
            if 2 <= len(name_cand) <= 8:
                candidates.append((name_cand, 2))

    # 模式4: snippet中包名前后的中文词（权重1）
    for text in snippets:
        if package_name not in text:
            continue
        idx = text.find(package_name)
        context_before = text[max(0, idx - 50): idx]
        m_before = re.search(r'([\u4e00-\u9fff]{2,10})\s*(?:是|即|为|[：:])\s*$', context_before)
        if m_before:
            candidates.append((m_before.group(1).strip(), 1))

    if not candidates:
        return None

    vote_counter = Counter()
    for name_cand, weight in candidates:
        vote_counter[name_cand] += weight

    _noise = {"下载", "安装", "官网", "应用", "软件", "版本", "更新", "手机", "最新", "免费", "正版",
              "谷歌", "官方", "独立版", "手表版", "旧版本", "安卓", "手游", "游戏", "客户端",
              "安装包", "汉化版", "中文版", "破解版", "精简版"}
    for noise in _noise:
        vote_counter.pop(noise, None)

    # 过滤描述性短语：以"一"开头（一款/一个/一种）或含有动词/形容词的非名称词
    # 正则已在模块级预编译为 _DESC_PHRASE_RE
    for name in list(vote_counter.keys()):
        if _DESC_PHRASE_RE.search(name):
            del vote_counter[name]

    if not vote_counter:
        return None

    best_name = vote_counter.most_common(1)[0][0]
    return best_name if 2 <= len(best_name) <= 15 else None


def _get_distinctive_pkg_keywords(package_name):
    """从包名中提取最独特的搜索关键词，过滤 android/com/net 等通用词。
    例：android.yoctsyo.sdjsql → ['yoctsyo', 'sdjsql']
        com.tencent.mm       → ['tencent', 'mm']
    """
    # 通用技术词，作为包名前缀/段时意义不大
    _common = {"android", "com", "net", "org", "io", "cn", "app", "apps",
               "mobile", "inc", "ltd", "game", "games", "studio", "studios",
               "tech", "dev", "official", "group", "co", "pub", "pro",
               "free", "plus", "lite", "hd", "vip"}
    parts = re.split(r'[\._\-]', package_name.lower())
    keywords = [p for p in parts if len(p) >= 3 and p not in _common]
    return keywords


def _search_toutiao_json(package_name, kw_query=None):
    """用头条搜索（so.toutiao.com）反查包名。
    页面以服务端渲染的方式将结果嵌入 JSON（title.text / abstract.text），
    无需 JS 渲染，且不易封锁（无 cookie/anti-spider 要求）。
    返回：(app_name_or_None, ref_url_or_None)
    """
    if kw_query is None:
        kw_query = package_name
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "zh-CN,zh;q=0.9",
    }
    # 优先用独特词查询，再用完整包名精确查询
    queries = [f"{kw_query} APK下载"]
    if kw_query != package_name:
        queries.append(f"{package_name} APK下载")

    all_titles = []
    all_snippets = []
    best_ref_url = None

    for q in queries:
        try:
            url = f"https://so.toutiao.com/search?keyword={urllib.parse.quote(q)}&pd=synthesis"
            resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
            if resp.status_code != 200 or len(resp.text) < 5000:
                continue
            raw = resp.content.decode("utf-8", errors="replace")
            # 提取 title.text 格式（头条搜索结果标题）
            titles = re.findall(r'"title"\s*:\s*\{[^}]{0,200}?"text"\s*:\s*"([^"]+)"', raw)
            for t in titles:
                # 去除 HTML 标签（<em>等高亮标签）
                t_clean = re.sub(r'<[^>]+>', '', t).strip()
                if t_clean and len(t_clean) >= 3:
                    all_titles.append(t_clean)
            # 提取所有 text 字段作为摘要备选
            all_raw_texts = re.findall(r'"text"\s*:\s*"([^"]{10,300})"', raw)
            for t in all_raw_texts:
                t_clean = re.sub(r'<[^>]+>', '', t).strip()
                if t_clean and package_name not in t_clean:
                    all_snippets.append(t_clean)
            # 提取参考链接
            if best_ref_url is None:
                raw_urls = re.findall(r'"url"\s*:\s*"(https?://[^"]{10,200})"', raw)
                for u in raw_urls:
                    if any(kw in u.lower() for kw in _PREFERRED_APK_DOMAINS):
                        best_ref_url = u
                        break
            if all_titles:
                break
        except Exception:
            continue

    if not all_titles and not all_snippets:
        return None, None

    app_name = _extract_app_name_from_titles(all_titles, all_snippets, package_name)
    return app_name, best_ref_url


def _search_sogou_web(package_name):
    """用搜狗网页搜索反查包名，SSR页面稳定可靠，不易封锁。
    依次尝试多种查询格式，从结果标题/摘要中投票返回应用名。
    返回：(app_name_or_None, ref_url_or_None)
    """
    # 限速：两次请求间隔至少3秒，避免触发搜狗反爬
    elapsed = time.time() - _sogou_search_last_call[0]
    if elapsed < 3.0:
        time.sleep(3.0 - elapsed)
    _sogou_search_last_call[0] = time.time()

    sess = _get_sogou_session()

    # 构建搜索查询：优先用独特词，避免 android/com 等通用前缀干扰搜索结果
    distinct_kws = _get_distinctive_pkg_keywords(package_name)
    kw_query = " ".join(distinct_kws) if distinct_kws else package_name

    # 先尝试 Toutiao（头条搜索）——数据量大、不易封锁、服务端渲染 JSON 嵌入页面
    app_name, ref_url = _search_toutiao_json(package_name, kw_query)
    if app_name:
        return app_name, ref_url

    # Toutiao 失败则用搜狗 SSR 兜底
    all_titles = []
    all_snippets = []
    best_ref_url = None
    query_formats = [f"{kw_query} 安卓 APK", f"{package_name} APK下载"]
    if kw_query == package_name:
        query_formats = [f"{package_name} APK下载", f"{package_name} 安卓"]

    for q in query_formats:
        try:
            url = f"https://www.sogou.com/web?query={urllib.parse.quote(q)}"
            resp = sess.get(url, timeout=HTTP_TIMEOUT)
            if resp.status_code != 200:
                continue
            # 检测反爬验证页（antispider.min.css 是反爬标志）
            if "antispider" in resp.text or len(resp.text) < 10000:
                global _sogou_session
                _sogou_session = None
                continue
            soup = BeautifulSoup(resp.text, _BS_PARSER)
            for h3 in soup.select("h3 a"):
                title = h3.get_text(strip=True)
                if title:
                    all_titles.append(title)
            for sel in [".fz-mid", ".str_info", ".space-txt", "p.str-short"]:
                for elem in soup.select(sel):
                    text = elem.get_text(strip=True)
                    if text and len(text) > 10:
                        all_snippets.append(text)
            # 注意：citeLinkClass span 显示的是截断的展示URL（如 https://xxx.com/a...），
            # 不是真实完整链接，不能使用。Sogou 不提取 ref_url。
            if all_titles:
                break
        except Exception:
            continue

    if not all_titles and not all_snippets:
        return None, None

    app_name = _extract_app_name_from_titles(all_titles, all_snippets, package_name)
    return app_name, best_ref_url


def _search_360(package_name):
    """用360搜索（so.com）反查包名，作为备用引擎。
    有IP封锁风险，加入3秒限速和封锁检测。
    返回：(app_name_or_None, ref_url_or_None)
    """
    # 限速：两次请求间隔至少3秒
    elapsed = time.time() - _360_search_last_call[0]
    if elapsed < 3.0:
        time.sleep(3.0 - elapsed)
    _360_search_last_call[0] = time.time()

    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    try:
        distinct_kws = _get_distinctive_pkg_keywords(package_name)
        kw_q = " ".join(distinct_kws) if distinct_kws else package_name
        url = f"https://www.so.com/s?q={urllib.parse.quote(kw_q + ' 安卓 APK')}"
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200 or len(resp.text) < 2000:
            return None, None
        # 检测360限流/验证码页面
        if "访问异常" in resp.text or "verify" in resp.url or len(resp.text) < 8000:
            return None, None

        soup = BeautifulSoup(resp.text, _BS_PARSER)
        titles, snippets, ref_urls = [], [], []
        for item in soup.select("li.res-list"):
            t = item.select_one("h3.res-title a")
            s = item.select_one("span.res-list-summary")
            if t:
                titles.append(t.get_text(strip=True))
                mdurl = t.get("data-mdurl", "")
                if mdurl:
                    ref_urls.append(mdurl)
            if s:
                snippets.append(s.get_text(strip=True))

        if not titles and not snippets:
            return None, None

        # 选最优参考链接
        best_ref_url = None
        for u in ref_urls:
            if any(kw in u.lower() for kw in _PREFERRED_APK_DOMAINS):
                best_ref_url = u
                break
        if not best_ref_url and ref_urls:
            best_ref_url = ref_urls[0]

        app_name = _extract_app_name_from_titles(titles, snippets, package_name)
        return app_name, best_ref_url

    except Exception:
        return None, None


def search_360_for_app_name(package_name):
    """搜索引擎反查包名对应的应用名称和参考链接（兜底策略）。
    搜索顺序：头条搜索（最稳定）→ 搜狗SSR → 360搜索（易封锁，最后备用）
    返回：(app_name, ref_url) 元组，均可能为 None
    """
    with _search_engine_cache_lock:
        cached = _search_engine_cache.get(package_name)
    if cached is not None:
        return cached

    distinct_kws = _get_distinctive_pkg_keywords(package_name)
    kw_query = " ".join(distinct_kws) if distinct_kws else package_name

    # 第一步：头条搜索（Toutiao）——JSON内嵌数据，稳定，不封锁
    app_name, ref_url = _search_toutiao_json(package_name, kw_query)

    # 第二步：搜狗 SSR 搜索
    if not app_name:
        app_name, ref_url_sogou = _search_sogou_web(package_name)
        if not ref_url:
            ref_url = ref_url_sogou

    # 第三步：360搜索（备用，易封锁）
    if not app_name:
        app_name, ref_url_360 = _search_360(package_name)
        if not ref_url:
            ref_url = ref_url_360

    result = (app_name, ref_url)
    with _search_engine_cache_lock:
        _search_engine_cache[package_name] = result
    return result


def _fetch_360_store_detail(app_id, headers=None):
    """抓取360应用商店详情页（m.app.so.com/detail/index?id=），返回应用信息dict
    可提取字段：包名、应用名、下载链接、图标、版本、大小、评分、开发者
    注意：分类字段页面内不含（动态加载），始终返回空字符串
    """
    if headers is None:
        headers = {
            "User-Agent": "Mozilla/5.0 (Linux; Android 12; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9",
        }
    try:
        url = f"https://m.app.so.com/detail/index?id={app_id}"
        resp = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200 or len(resp.text) < 2000:
            return None
        soup = BeautifulSoup(resp.text, _BS_PARSER)
        body_text = soup.get_text()

        # 应用名
        name_el = soup.select_one(".app-name")
        app_name = clean_app_name(name_el.get_text(strip=True)) if name_el else ""

        # 下载链接 & 图标（在 data-url / data-logo 属性上）
        dl_el = soup.select_one("[data-url]")
        download_url = dl_el["data-url"] if dl_el else ""
        icon_url = dl_el.get("data-logo", "") if dl_el else ""

        # 包名在页面内嵌 JS 的 appInfo 块里
        package_name = ""
        for script in soup.find_all("script"):
            c = script.string or ""
            if "'apkid'" in c:
                m = re.search(r"'apkid'\s*:\s*'([^']+)'", c)
                if m:
                    package_name = m.group(1)
                    break

        # 版本（在 <p> 标签文本中：版本：x.x.x）
        version_m = re.search(r'版本[：:]\s*([\d.A-Za-z]+)', body_text)
        version = version_m.group(1) if version_m else ""

        # 大小（如 56.36Mb）
        size_m = re.search(r'([\d.]+)\s*Mb', body_text, re.I)
        size = size_m.group(1) + "MB" if size_m else ""

        # 评分
        score_m = re.search(r'([\d.]+)分', body_text)
        score = score_m.group(1) if score_m else ""

        # 开发者
        dev_m = re.search(r'开发[者商]?[：:]\s*([^\n]{2,40})', body_text)
        developer = dev_m.group(1).strip() if dev_m else ""

        if not app_name or not package_name:
            return None

        # download_url 使用商店页面地址（稳定、无时效）
        # CDN 直链（download_url 原始值）仅作为 apk_direct_url 保存，供 APK 下载功能使用
        store_page_url = f"https://m.app.so.com/detail/index?id={app_id}"

        return {
            "package_name":    package_name,
            "platform":        "Android",
            "app_name":        app_name,
            "download_url":    store_page_url,   # 商店页面，稳定不过期
            "apk_direct_url":  download_url,     # CDN直链，有时效token，仅供APK下载
            "icon_url":        icon_url,
            "category":        "",               # 360商店详情页分类为动态加载，无法从SSR获取
            "version":         version,
            "size":            size,
            "score":           score,
            "developer":       developer,
            "source":          "360应用商店",
        }
    except Exception:
        return None


def _bing_mobile_search(query):
    """用必应移动端搜索，返回 HTML 文本（失败返回空字符串）。
    实测：iPhone/Android/WeChat 三种移动 UA 都能触发 rich-card，
    其他搜索引擎（百度/搜狗/360移动）均被验证码封锁。
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        "Accept-Language": "zh-CN,zh;q=0.9",
    }
    try:
        url = f"https://cn.bing.com/search?q={urllib.parse.quote(query)}&mkt=zh-CN"
        r = _HTTP.get(url, headers=headers, timeout=HTTP_TIMEOUT)
        if r.status_code == 200 and len(r.text) > 5000:
            return r.text
    except Exception:
        pass
    return ""


def search_stores_via_bing(app_name, expected_package=None):
    """用必应移动端搜索应用名，从结果中提取各商店链接并验证包名。
    必应移动端会展示各商店的 rich-card，包含直接商店链接。
    尝试多种查询组合，覆盖 360/腾讯/小米/豌豆荚 等主流商店。
    返回第一个包名匹配的商店结果 dict，找不到返回 None。
    """
    # 多种查询组合（按命中率排序）
    queries = [
        f"{app_name} 360应用",
        f"{app_name} 应用宝",
        f"{app_name} 安卓 APK",
        f"{app_name} 小米应用商店",
        f"{app_name} 下载",
    ]
    # 各商店 URL 的正则和对应 fetch 函数
    store_patterns = [
        # (正则, 提取组, fetch函数或处理方式, 商店名)
        (r'app\.so\.com/detail/index\?id=(\d+)',               1, "360",       "360应用商店"),
        (r'sj\.qq\.com/appdetail/([\w\.]+)',                   1, "tencent",   "腾讯应用宝"),
        (r'app\.mi\.com/details\?id=([\w\.]+)',                1, "xiaomi",    "小米应用商店"),
        (r'www\.wandoujia\.com/apps/([\w\.]+)',                 1, "wandoujia", "豌豆荚"),
        (r'app\.flyme\.cn/apps/public/detail\?package_name=([\w\.]+)', 1, "flyme", "魅族Flyme"),
        (r'www\.appchina\.com/app/([\w\.]+)',                  1, "appchina",  "应用汇"),
    ]

    collected = {}   # store_key → [id_or_pkg, ...]

    for q in queries:
        html = _bing_mobile_search(q)
        if not html:
            continue
        for pattern, grp, store_key, _ in store_patterns:
            found = re.findall(pattern, html)
            if found:
                if store_key not in collected:
                    collected[store_key] = []
                for f in found:
                    if f not in collected[store_key]:
                        collected[store_key].append(f)
        # 已有至少一个商店命中就不再继续（避免多余请求）
        if collected:
            break

    if not collected:
        return None

    # 按优先级尝试各商店（优先数据最完整的）
    priority = ["360", "tencent", "xiaomi", "wandoujia", "flyme", "appchina"]
    store_fetch_map = {
        "360":       lambda sid:  _fetch_360_store_detail(sid),
        "tencent":   lambda pkg:  search_tencent(pkg),
        "xiaomi":    lambda pkg:  search_xiaomi(pkg),
        "wandoujia": lambda pkg:  search_wandoujia(pkg),
        "flyme":     lambda pkg:  search_flyme(pkg),
        "appchina":  lambda pkg:  search_appchina(pkg),
    }

    for store_key in priority:
        if store_key not in collected:
            continue
        fetch_fn = store_fetch_map.get(store_key)
        if not fetch_fn:
            continue
        for id_or_pkg in collected[store_key]:
            try:
                result = fetch_fn(id_or_pkg)
            except Exception:
                continue
            if not result:
                continue
            # 验证包名
            if expected_package and result.get("package_name") != expected_package:
                continue
            return result

    return None


def search_360_store_by_app_name(app_name, expected_package=None):
    """通过必应移动端搜索定位360应用商店条目（或其他商店），获取完整应用信息。
    内部调用 search_stores_via_bing，若未命中则单独重试360商店查询。
    """
    # 先尝试多商店扫描
    result = search_stores_via_bing(app_name, expected_package)
    if result:
        return result

    # 多商店未命中，最后单独用包名直接查360（有时应用名查不中但包名可以）
    if expected_package:
        try:
            html = _bing_mobile_search(f"{expected_package} 360应用")
            if html:
                ids = re.findall(r'app\.so\.com/detail/index\?id=(\d+)', html)
                for app_id in dict.fromkeys(ids):
                    r = _fetch_360_store_detail(app_id)
                    if r and r.get("package_name") == expected_package:
                        return r
        except Exception:
            pass

    return None


def search_bing_for_android_package(name, limit=5):
    """用 Bing 搜索 app 名称，从应用商店结果 URL 中提取 Android 包名。
    适用于国内网络环境，Bing 可访问且搜索结果常包含应用商店链接。
    例：搜索"豆包"可从 sj.qq.com/appdetail/com.larus.nova 提取包名。
    """
    query = f'"{name}" site:sj.qq.com OR site:app.mi.com OR site:www.wandoujia.com OR site:www.appchina.com'
    url = f"https://www.bing.com/search?q={urllib.parse.quote(query)}&setlang=zh-CN&cc=CN"
    try:
        bing_headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        }
        resp = _HTTP.get(url, headers=bing_headers, timeout=5)
        if resp.status_code != 200:
            return []
        # 从应用商店 URL 中提取包名
        # 腾讯应用宝: sj.qq.com/appdetail/PACKAGE  或  sj.qq.com/detail/.../PACKAGE
        # 小米: app.mi.com/details?id=PACKAGE
        # 豌豆荚: www.wandoujia.com/apps/PACKAGE
        # 应用汇: www.appchina.com/app/PACKAGE
        patterns = [
            r'sj\.qq\.com/appdetail/([\w\.]+)',
            r'sj\.qq\.com/detail/[^/]+/([\w\.]+)',
            r'app\.mi\.com/details\?id=([\w\.]+)',
            r'wandoujia\.com/apps/([\w\.]+)',
            r'appchina\.com/app/([\w\.]+)',
        ]
        noise_prefixes = (
            "com.google.", "com.android.", "com.sec.", "com.samsung.",
            "com.miui.", "com.huawei.", "com.xiaomi.", "com.tencent.mobileqq",
        )
        seen, result = set(), []
        for pat in patterns:
            for pkg in re.findall(pat, resp.text):
                if pkg in seen:
                    continue
                seen.add(pkg)
                if any(pkg.startswith(px) for px in noise_prefixes):
                    continue
                result.append(pkg)
                if len(result) >= limit:
                    return result
        return result
    except Exception:
        return []



def query_single(package_name, android_store_order=None, timeout_override=None):
    """查询单个包名，返回结果列表。timeout_override 用于重查时延长超时。"""
    package_name = clean_package_name(package_name)
    _timeout = timeout_override or HTTP_TIMEOUT
    if not package_name:
        return []

    if android_store_order is None:
        android_store_order = get_ranked_store_order()

    # ── 第1层：所有商店 + Apple 并发（复用共享线程池） ──────────────────
    # 注意：不再 new ThreadPoolExecutor，批量场景下省去 50 次池创建/销毁开销
    android_futures = {}
    _store_start_times = {}  # 记录每个商店查询开始时间
    for store_id in android_store_order:
        func = STORE_SEARCH_FUNCS.get(store_id)
        if func:
            _store_start_times[store_id] = time.time()
            android_futures[store_id] = _QUERY_POOL.submit(func, package_name)
    future_apple = _QUERY_POOL.submit(search_apple, package_name)

    # Flyme / 头条 / 360 / 搜狗 全部"懒触发"：只有主商店漏掉时才启动
    # 原来 Flyme 是投机预取，但主商店命中率 >90%，95% 的预取都被丢弃——纯浪费
    future_flyme = None
    distinct_kws = _get_distinctive_pkg_keywords(package_name)
    kw_query = " ".join(distinct_kws) if distinct_kws else package_name
    future_toutiao = None  # 延迟创建，仅在需要时提交到 _QUERY_POOL

    # "先到先用"：按 as_completed 收集结果，一旦有 Android 命中就等待高优先级商店
    android_result = None
    android_all_names = []
    _store_results = {}  # store_id → result

    # 先用 as_completed 快速收集已完成的结果（最多等 _timeout 秒）
    all_android_futures = {v: k for k, v in android_futures.items()}
    all_futures = list(android_futures.values()) + [future_apple]
    _query_start_time = time.time()
    try:
        for f in concurrent.futures.as_completed(all_futures, timeout=_timeout):
            if f == future_apple:
                continue
            sid = all_android_futures.get(f, "")
            try:
                result = f.result(timeout=0)
            except Exception:
                result = None
            if result:
                _store_results[sid] = result
                android_all_names.append(result.get("app_name", ""))
                # 一旦有任何一个商店命中，等待高优先级商店直到 _timeout 截止
                # 防止低优先级商店抢先，导致高优先级商店（如小米）的结果被忽略
                if android_result is None:
                    android_result = result
                    # 计算剩余等待时间（总超时减去已用时间，至少1.5秒）
                    # 地板从 0.5s 提到 1.5s：Windows 上豌豆荚/应用汇常先返回，小米/应用宝
                    # 需要额外时间窗口才能赶上，否则低优先级商店会"抢先定案"导致图标缺失
                    _elapsed = time.time() - _query_start_time
                    _remaining = max(1.5, _timeout - _elapsed)
                    # 针对当前命中商店之前的高优先级商店，用剩余时间等待
                    for hi_sid in android_store_order:
                        if hi_sid == sid:
                            break  # 当前商店已是最高优先级，无需等待
                        if hi_sid in _store_results:
                            break  # 更高优先级商店已有结果
                        hi_f = android_futures.get(hi_sid)
                        if hi_f and not hi_f.done():
                            # 用剩余时间等待，地板 1.0s（Windows 下小米/应用宝常需要这个窗口）
                            _wait = max(1.0, _timeout - (time.time() - _query_start_time))
                            try:
                                hi_r = hi_f.result(timeout=_wait)
                                if hi_r:
                                    _store_results[hi_sid] = hi_r
                                    android_all_names.append(hi_r.get("app_name", ""))
                            except Exception:
                                pass
                    # 顺手收集其他已完成的 futures
                    for ff in all_futures:
                        if ff != future_apple and ff.done():
                            try:
                                extra_sid = all_android_futures.get(ff, "")
                                if extra_sid and extra_sid not in _store_results:
                                    er = ff.result(timeout=0)
                                    if er:
                                        _store_results[extra_sid] = er
                                        android_all_names.append(er.get("app_name", ""))
                            except Exception:
                                pass
                    break
    except concurrent.futures.TimeoutError:
        pass

    # 收集已完成但还没处理的 store 结果（名称优化用）
    for sid, f in android_futures.items():
        if sid not in _store_results and f.done():
            try:
                r = f.result(timeout=0)
                if r:
                    _store_results[sid] = r
                    android_all_names.append(r.get("app_name", ""))
            except Exception:
                pass

    # 记录商店查询统计（用于自动排名）
    _now = time.time()
    for sid in android_futures:
        elapsed = _now - _store_start_times.get(sid, _now)
        hit = sid in _store_results
        _record_store_stat(sid, hit, elapsed)

    # 按优先级选最佳 Android 结果（完整结果优先于不完整结果）
    # 完整 = app_name + icon_url + category + download_url 全部非空
    def _is_complete(r):
        return bool(r.get("app_name") and r.get("icon_url") and
                     r.get("category") and r.get("download_url"))

    if _store_results:
        # 第一轮：按优先级找完整结果
        best = None
        for sid in android_store_order:
            r = _store_results.get(sid)
            if r and _is_complete(r):
                best = r
                break
        # 第二轮：没有完整结果时，按优先级用不完整的
        if best is None:
            for sid in android_store_order:
                if sid in _store_results:
                    best = _store_results[sid]
                    break
        if best:
            android_result = best

        # 跨商店补全：用其他商店的 icon_url / category 填补当前结果的空缺
        if android_result:
            for sid in android_store_order:
                r = _store_results.get(sid)
                if not r or r is android_result:
                    continue
                if not android_result.get("icon_url") and r.get("icon_url"):
                    android_result["icon_url"] = r["icon_url"]
                if not android_result.get("category") and r.get("category"):
                    android_result["category"] = r["category"]
                if android_result.get("icon_url") and android_result.get("category"):
                    break  # 已补全

    # Apple 结果
    try:
        ios_result = future_apple.result(timeout=0.5) if not future_apple.done() else future_apple.result(timeout=0)
    except Exception:
        ios_result = None

    # 主商店全部未找到时：懒触发 Flyme（不在主列表时才需要）
    if android_result is None and "flyme" not in android_store_order:
        try:
            flyme_result = _QUERY_POOL.submit(search_flyme, package_name).result(
                timeout=_timeout
            )
            if flyme_result:
                android_result = flyme_result
        except Exception:
            pass

    # ── 第2层：搜索引擎反查（头条懒触发，之后搜狗+360并发） ────────────────
    if android_result is None:
        # 懒触发头条搜索（到这里主商店 + flyme 全漏，才值得跑搜索引擎）
        if future_toutiao is None:
            future_toutiao = _QUERY_POOL.submit(_search_toutiao_json, package_name, kw_query)
        toutiao_name, toutiao_ref = future_toutiao.result()
        app_name_hint = toutiao_name
        ref_url = toutiao_ref

        # 头条没找到 → 搜狗 + 360 并发（复用共享池）
        if not app_name_hint:
            f_sogou = _QUERY_POOL.submit(_search_sogou_web, package_name)
            f_360   = _QUERY_POOL.submit(_search_360, package_name)
            sogou_name, sogou_ref = f_sogou.result()
            s360_name, s360_ref   = f_360.result()
            app_name_hint = sogou_name or s360_name
            ref_url = ref_url or sogou_ref or s360_ref

        # 缓存搜索引擎结果
        with _search_engine_cache_lock:
            _search_engine_cache[package_name] = (app_name_hint, ref_url)

        if app_name_hint:
            android_all_names.append(app_name_hint)

            # ── 第3层：用名称回查各商店（共享池并发） ────────
            name_search_futures = {}
            for store_id in android_store_order:
                name_func = STORE_NAME_SEARCH_FUNCS.get(store_id)
                if name_func:
                    name_search_futures[store_id] = _QUERY_POOL.submit(
                        name_func, app_name_hint)
            # 同时发起 Bing 多商店搜索
            f_bing_stores = _QUERY_POOL.submit(
                search_360_store_by_app_name, app_name_hint, package_name)

            # 按优先级检查名称搜索结果
            for store_id in android_store_order:
                if store_id not in name_search_futures:
                    continue
                try:
                    store_results = name_search_futures[store_id].result() or []
                    for r in store_results:
                        if r.get("package_name") == package_name:
                            android_result = r
                            break
                except Exception:
                    continue
                if android_result:
                    break

            # 商店名称搜索没命中 → 用 Bing 多商店结果
            if android_result is None:
                so_result = f_bing_stores.result()
                if so_result:
                    android_result = so_result

            # 仍找不到，返回名称 + 参考链接；无参考链接时用七麦搜索页兜底
            if android_result is None:
                qimai_fallback = f"https://www.qimai.cn/search/android/search/{urllib.parse.quote(package_name)}"
                android_result = {
                    "package_name": package_name,
                    "platform":     "Android",
                    "app_name":     app_name_hint,
                    "download_url": ref_url or qimai_fallback,
                    "icon_url":     "",
                    "category":     "",
                    "source":       "search_engine_ref" if ref_url else "qimai_hint",
                }

    rows = []

    if android_result:
        # 以按优先级选出的 android_result.app_name 作为基准名，避免"首响应的低优先级商店"
        # 把无关 App（如 com.qiekj.user 匹配到"三国"）带成基准名
        best_name = pick_best_name(android_all_names, primary=android_result.get("app_name")) \
            or android_result.get("app_name", "")
        _icon = android_result.get("icon_url", "")
        _cat  = android_result.get("category", "")
        rows.append(_mark_incomplete({
            "package_name": package_name,
            "platform": "Android",
            "app_name": best_name,
            "download_url": android_result["download_url"],
            "icon_url": _icon,
            "category": _cat,
            "source": android_result.get("source", ""),
        }))

    if ios_result:
        rows.append(_mark_incomplete({
            "package_name": package_name,
            "platform": "iOS",
            "app_name": ios_result["app_name"],
            "download_url": ios_result["download_url"],
            "icon_url": ios_result.get("icon_url", ""),
            "category": ios_result.get("category", ""),
        }))

    if not rows:
        rows.append(_mark_incomplete({
            "package_name": package_name,
            "platform":     "Android",
            "app_name":     "未找到",
            "download_url": f"https://www.qimai.cn/search/android/search/{urllib.parse.quote(package_name)}",
            "icon_url":     "",
            "category":     "",
            "source":       "qimai_hint",
        }))

    return rows


def _enrich_rows(rows, get_apk_url, apk_url_mode, get_sha1, get_sha256=False):
    """为结果列表中的每个 Android 行补充 APK 直链、SHA1、SHA256。
    批量场景下并发查 appchina + wandoujia，不再逐条串行。"""
    need_appchina = get_apk_url or get_sha1 or get_sha256
    # 收集需要补齐的 Android pkg
    android_rows = [r for r in rows
                    if r.get("platform") == "Android" and r.get("package_name")]
    if not android_rows:
        return rows
    pkgs = list({r["package_name"] for r in android_rows})

    # 并发跑 appchina（供 APK 直链 + SHA 共用）
    ac_map = {}  # pkg -> apk_direct_url
    if need_appchina:
        ac_results = _parallel_map(search_appchina, pkgs, timeout=15)
        for pkg, ac in ac_results.items():
            ac_map[pkg] = (ac or {}).get("apk_direct_url", "")

    # 并发跑 wandoujia（仅 multiple 模式需要）
    wdj_map = {}  # pkg -> apk_direct_url
    if get_apk_url and apk_url_mode == "multiple":
        wdj_results = _parallel_map(_get_wandoujia_apk_url, pkgs, timeout=15)
        wdj_map = {k: (v or "") for k, v in wdj_results.items()}

    # 把并发结果贴回每一行
    for r in android_rows:
        pkg = r["package_name"]
        apk_url_from_appchina = ac_map.get(pkg, "")
        if get_apk_url:
            apk_urls = []
            if apk_url_from_appchina:
                apk_urls.append(apk_url_from_appchina)
            wdj_url = wdj_map.get(pkg, "")
            if wdj_url and wdj_url not in apk_urls:
                apk_urls.append(wdj_url)
            r["apk_direct_urls"] = apk_urls

    # SHA1/SHA256：每行的 APK 下载本身不能在进程级去重（不同 pkg 对应不同 APK）
    # 但不同行对应同一 pkg 时可以共用，按 sha_src 去重
    if get_sha1 or get_sha256:
        # 收集所有需要取指纹的 apk_url
        src_map = {}  # row -> sha_src url
        for r in android_rows:
            pkg = r["package_name"]
            sha_src = ac_map.get(pkg, "")
            if not sha_src and get_apk_url:
                apk_urls = r.get("apk_direct_urls") or []
                sha_src = apk_urls[0] if apk_urls else ""
            src_map[id(r)] = sha_src
        unique_srcs = list({s for s in src_map.values() if s})
        fp_cache = _parallel_map(_extract_cert_fingerprints, unique_srcs, timeout=60) \
            if unique_srcs else {}
        for r in android_rows:
            sha_src = src_map.get(id(r), "")
            fp = fp_cache.get(sha_src) if sha_src else None
            sha1, sha256 = fp if fp else ("", "")
            if get_sha1:
                r["sha1"] = sha1
            if get_sha256:
                r["sha256"] = sha256

    return rows


def query_single_extended(package_name, android_store_order=None,
                          get_apk_url=False, apk_url_mode="single",
                          get_sha1=False, get_sha256=False,
                          timeout_override=None):
    """查询单个包名，含可选的 APK 直链、SHA1、SHA256"""
    rows = query_single(package_name, android_store_order, timeout_override=timeout_override)
    return _enrich_rows(rows, get_apk_url, apk_url_mode, get_sha1, get_sha256)


def query_by_name_extended(name, android_store_order=None, exact_search=False,
                           get_apk_url=False, apk_url_mode="single",
                           get_sha1=False, get_sha256=False):
    """按名称查询，含可选的 APK 直链、SHA1、SHA256"""
    rows = query_by_name(name, android_store_order, exact_search)
    return _enrich_rows(rows, get_apk_url, apk_url_mode, get_sha1, get_sha256)


# ========== 路由 ==========

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/retry", methods=["POST"])
def api_retry():
    """重查不完整结果：超时加倍，并行所有方法"""
    req_data = request.get_json()
    package_names = req_data.get("package_names", [])
    if not package_names:
        return jsonify({"results": []})

    retry_timeout = HTTP_TIMEOUT * 2  # 重查时超时翻倍

    def _retry_one(pkg):
        try:
            return query_single(pkg, timeout_override=retry_timeout)
        except Exception:
            return [_mark_incomplete({
                "package_name": pkg, "platform": "Android",
                "app_name": "未找到",
                "download_url": f"https://www.qimai.cn/search/android/search/{urllib.parse.quote(pkg)}",
                "icon_url": "", "category": "",
                "source": "qimai_hint",
            })]

    all_results = []
    with concurrent.futures.ThreadPoolExecutor(
            max_workers=min(20, len(package_names))) as exe:
        futures = {exe.submit(_retry_one, pkg): pkg for pkg in package_names}
        for f in concurrent.futures.as_completed(futures):
            try:
                rows = f.result()
                all_results.extend(rows or [])
            except Exception:
                pkg = futures[f]
                all_results.append(_mark_incomplete({
                    "package_name": pkg, "platform": "Android",
                    "app_name": "未找到",
                    "download_url": f"https://www.qimai.cn/search/android/search/{urllib.parse.quote(pkg)}",
                    "icon_url": "", "category": "",
                    "source": "qimai_hint",
                }))

    return jsonify({"results": all_results})


def _safe_filename(s):
    """把任意字符串处理成对文件系统安全的文件名（保留中英文，替换特殊符号）"""
    if not s:
        return "icon"
    # Windows / *nix 都不允许这些字符
    s = re.sub(r'[\\/:*?"<>|\r\n\t]', "_", s)
    # 折叠多空格、去首尾空白和点（Windows 不允许结尾的 . 和空格）
    s = re.sub(r"\s+", " ", s).strip().strip(".")
    if not s:
        return "icon"
    # 太长会被某些文件系统截断，限制一下
    if len(s) > 80:
        s = s[:80].rstrip()
    return s


def _ext_from_content_type(ct):
    if not ct:
        return ""
    ct = ct.lower().split(";")[0].strip()
    mapping = {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "image/bmp": ".bmp",
        "image/x-icon": ".ico",
        "image/vnd.microsoft.icon": ".ico",
        "image/svg+xml": ".svg",
    }
    return mapping.get(ct, "")


def _ext_from_url(url):
    try:
        path = urllib.parse.urlparse(url).path.lower()
        m = re.search(r"\.(png|jpg|jpeg|webp|gif|bmp|ico|svg)(?:$|[?#])", path)
        if m:
            ext = "." + m.group(1)
            return ".jpg" if ext == ".jpeg" else ext
    except Exception:
        pass
    return ""


def _fetch_icon_bytes(url, timeout=8):
    """拉取图标，返回 (content_bytes, ext, content_type)。失败返回 (None, '', '')"""
    if not url:
        return None, "", ""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "image/*,*/*;q=0.8",
        }
        resp = _HTTP.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        if resp.status_code != 200 or not resp.content:
            return None, "", ""
        ct = resp.headers.get("Content-Type", "")
        ext = _ext_from_content_type(ct) or _ext_from_url(url) or ".png"
        return resp.content, ext, ct
    except Exception:
        return None, "", ""


def _content_disposition(filename):
    """生成带 UTF-8 编码的 Content-Disposition（兼容中文文件名）"""
    # ASCII fallback 去掉非 ASCII 字符
    ascii_name = re.sub(r"[^\x20-\x7e]", "_", filename) or "icon"
    quoted = urllib.parse.quote(filename, safe="")
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quoted}"


@app.route("/api/icon", methods=["GET"])
def api_icon():
    """下载单个 icon。参数：url（必填）、name（app 名）、platform"""
    url = request.args.get("url", "").strip()
    app_name = request.args.get("name", "").strip() or "icon"
    platform = request.args.get("platform", "").strip()

    if not url:
        return jsonify({"ok": False, "message": "missing url"}), 400

    content, ext, ct = _fetch_icon_bytes(url)
    if content is None:
        return jsonify({"ok": False, "message": "fetch failed"}), 502

    base = f"{app_name}-{platform}" if platform else app_name
    filename = _safe_filename(base) + ext
    resp = Response(content, mimetype=ct or "application/octet-stream")
    resp.headers["Content-Disposition"] = _content_disposition(filename)
    resp.headers["Content-Length"] = str(len(content))
    return resp


@app.route("/api/icons_zip", methods=["POST"])
def api_icons_zip():
    """批量下载 icons，打成 zip。body: {items: [{url, app_name, platform}, ...]}"""
    data = request.get_json(silent=True) or {}
    items = data.get("items", []) or []
    if not items:
        return jsonify({"ok": False, "message": "no items"}), 400
    if len(items) > 500:
        return jsonify({"ok": False, "message": "too many items"}), 400

    def _fetch(item):
        url = (item.get("url") or "").strip()
        app_name = (item.get("app_name") or "").strip() or "icon"
        platform = (item.get("platform") or "").strip()
        content, ext, _ct = _fetch_icon_bytes(url)
        return {
            "app_name": app_name,
            "platform": platform,
            "content": content,
            "ext": ext,
        }

    fetched = []
    with concurrent.futures.ThreadPoolExecutor(
            max_workers=min(8, len(items))) as exe:
        for res in exe.map(_fetch, items):
            fetched.append(res)

    # 打 zip（在内存中）
    buf = io.BytesIO()
    name_count = {}
    ok_count = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in fetched:
            if not r["content"]:
                continue
            base = f"{r['app_name']}-{r['platform']}" if r["platform"] else r["app_name"]
            safe = _safe_filename(base)
            fname = safe + (r["ext"] or ".png")
            # 去重：同名追加 -2, -3 ...
            if fname in name_count:
                name_count[fname] += 1
                stem, ext = os.path.splitext(fname)
                fname = f"{stem}-{name_count[fname]}{ext}"
            else:
                name_count[fname] = 1
            zf.writestr(fname, r["content"])
            ok_count += 1

    if ok_count == 0:
        return jsonify({"ok": False, "message": "all icons failed"}), 502

    buf.seek(0)
    date_str = time.strftime("%Y%m%d")
    zip_name = f"icons-{date_str}-{ok_count}.zip"
    resp = Response(buf.getvalue(), mimetype="application/zip")
    resp.headers["Content-Disposition"] = _content_disposition(zip_name)
    resp.headers["Content-Length"] = str(buf.getbuffer().nbytes)
    resp.headers["X-Icons-Count"] = str(ok_count)
    return resp


@app.route("/api/cancel_job/<job_id>", methods=["POST"])
def api_cancel_job(job_id):
    """取消指定 job，让 worker 尽早停止"""
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if job and job.get('status') == 'running':
            job['cancelled'] = True
    return jsonify({"ok": True})


@app.route("/api/shutdown", methods=["POST"])
def api_shutdown():
    """关闭服务。仅本机管理员可操作。"""
    if not _is_local_request_ip(request.remote_addr or ""):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    import threading
    def _stop():
        time.sleep(0.5)
        os._exit(0)
    threading.Thread(target=_stop, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/startup/status", methods=["GET"])
def api_startup_status():
    """查询开机自启状态"""
    enabled = _check_startup_enabled()
    return jsonify({"enabled": enabled})


@app.route("/api/lan_info", methods=["GET"])
def api_lan_info():
    """LAN 分享信息：开关状态、URL、二维码。
    二维码用 qrcode 库在服务端生成，base64 嵌入返回。"""
    try:
        import qrcode
        import base64
        port_here = int(os.environ.get("PORT", 9527))
        lan_ip = _get_lan_ip()
        is_lan_accessible = _is_real_lan_ip(lan_ip)
        url = f"http://{lan_ip}:{port_here}" if is_lan_accessible else ""

        qr_data_url = ""
        if is_lan_accessible:
            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10,
                border=2,
            )
            qr.add_data(url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="#1a1a1a", back_color="#ffffff")
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            qr_data_url = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")

        return jsonify({
            "url": url,
            "lan_ip": lan_ip,
            "port": port_here,
            "accessible": is_lan_accessible,
            "enabled": _LAN_ENABLED,
            "qr_data_url": qr_data_url,
            # 当前请求者是否是本机管理员（用于前端隐藏管理员专属按钮）
            "is_admin": _is_local_request_ip(request.remote_addr or ""),
        })
    except Exception as e:
        return jsonify({
            "url": "", "lan_ip": "", "port": 0,
            "accessible": False, "enabled": _LAN_ENABLED, "qr_data_url": "",
            "is_admin": _is_local_request_ip(request.remote_addr or ""),
            "error": str(e),
        })


@app.route("/api/history", methods=["GET"])
def api_history_get():
    """获取共享查询历史（所有客户端共用，方便手机等新设备看到本机的查询记录）"""
    with _HISTORY_LOCK:
        return jsonify({"history": list(_HISTORY)})


@app.route("/api/history", methods=["POST"])
def api_history_add():
    """添加一条历史记录。body: {entry: {...}}。
    本机和 LAN 访客都能添加，这样多设备查询都会进到同一份历史里。"""
    data = request.get_json(silent=True) or {}
    entry = data.get("entry")
    if not isinstance(entry, dict):
        return jsonify({"ok": False, "error": "missing entry"})
    # 瘦身：去掉过大的字段（图标 base64 等极端情况），保守策略
    results = entry.get("results") or []
    if isinstance(results, list):
        entry["results"] = results[:60]  # 单条最多 60 个结果
    with _HISTORY_LOCK:
        # 按时间戳去重（前端会带 timestamp）
        ts = entry.get("timestamp")
        _HISTORY[:] = [h for h in _HISTORY if h.get("timestamp") != ts]
        _HISTORY.insert(0, entry)
        del _HISTORY[_HISTORY_MAX:]
    _save_history()
    return jsonify({"ok": True, "count": len(_HISTORY)})


@app.route("/api/history", methods=["DELETE"])
def api_history_clear():
    """清空共享历史。只有本机管理员能清——LAN 访客不应该能擦掉本机的数据。"""
    if not _is_local_request_ip(request.remote_addr or ""):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    with _HISTORY_LOCK:
        _HISTORY.clear()
    _save_history()
    return jsonify({"ok": True})


@app.route("/api/about_info", methods=["GET"])
def api_about_info():
    """"关于"面板用的信息：作者微信二维码（服务端生成，不用打包图片）"""
    try:
        import qrcode
        import base64
        wechat_id = "rickaruike"
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2,
        )
        qr.add_data(wechat_id)
        qr.make(fit=True)
        img = qr.make_image(fill_color="#6A4FE5", back_color="#ffffff")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        qr_data_url = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")
        return jsonify({
            "wechat_id": wechat_id,
            "wechat_qr_data_url": qr_data_url,
        })
    except Exception as e:
        return jsonify({"wechat_id": "rickaruike", "wechat_qr_data_url": "", "error": str(e)})


@app.route("/api/lan_toggle", methods=["POST"])
def api_lan_toggle():
    """切换 LAN 开关。**仅限本机管理员操作**——非本机请求一律拒绝，
    否则 LAN 访客可以自行关闭共享（甚至把自己踢下线）。"""
    global _LAN_ENABLED
    client_ip = request.remote_addr or ""
    if not _is_local_request_ip(client_ip):
        return jsonify({
            "ok": False,
            "error": "forbidden",
            "message": "只有运行本工具的电脑管理员可以切换开关",
        }), 403
    try:
        data = request.get_json(silent=True) or {}
        enable = bool(data.get("enabled", False))
        _LAN_ENABLED = enable
        _save_lan_settings()
        if not enable:
            with _LAN_STATS_LOCK:
                _LAN_STATS["devices"].clear()
                _LAN_STATS["total_requests"] = 0
        return jsonify({"ok": True, "enabled": _LAN_ENABLED})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/lan_device_note", methods=["POST"])
def api_lan_device_note():
    """给一个 IP 设置/清空备注。仅本机管理员可操作。"""
    if not _is_local_request_ip(request.remote_addr or ""):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    data = request.get_json(silent=True) or {}
    ip = str(data.get("ip", "")).strip()
    note = str(data.get("note", "")).strip()[:40]  # 备注最长 40 字符
    if not ip:
        return jsonify({"ok": False, "error": "missing ip"})
    with _LAN_STATS_LOCK:
        if note:
            _LAN_DEVICE_NOTES[ip] = note
        else:
            _LAN_DEVICE_NOTES.pop(ip, None)
    _save_lan_settings()
    return jsonify({"ok": True, "ip": ip, "note": note})


@app.route("/api/lan_device_block", methods=["POST"])
def api_lan_device_block():
    """屏蔽/解除屏蔽一个 IP。仅本机管理员可操作。"""
    if not _is_local_request_ip(request.remote_addr or ""):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    data = request.get_json(silent=True) or {}
    ip = str(data.get("ip", "")).strip()
    blocked = bool(data.get("blocked", True))
    if not ip:
        return jsonify({"ok": False, "error": "missing ip"})
    with _LAN_STATS_LOCK:
        if blocked:
            _LAN_BLOCKED_IPS.add(ip)
        else:
            _LAN_BLOCKED_IPS.discard(ip)
    _save_lan_settings()
    return jsonify({"ok": True, "ip": ip, "blocked": blocked})


@app.route("/api/lan_stats", methods=["GET"])
def api_lan_stats():
    """LAN 访问统计：连接过的设备 + 每个设备的请求次数、最后活跃时间"""
    with _LAN_STATS_LOCK:
        now = time.time()
        devices = []
        for ip, d in _LAN_STATS["devices"].items():
            seconds_ago = int(now - d["last_seen"])
            devices.append({
                "ip": ip,
                "query_count": d.get("query_count", 0),
                "last_seen_ago_sec": seconds_ago,
                "first_seen_sec_ago": int(now - d["first_seen"]),
                "last_path": d.get("last_path", ""),
                "ua_short": _short_ua(d.get("ua", "")),
                "mac": d.get("mac", ""),
                "hostname": d.get("hostname", ""),
                "note": _LAN_DEVICE_NOTES.get(ip, ""),
                "blocked": ip in _LAN_BLOCKED_IPS,
            })
        # 把"有记录但已屏蔽"的设备也一并展示（即使它没在线）
        for ip in _LAN_BLOCKED_IPS:
            if ip not in _LAN_STATS["devices"]:
                devices.append({
                    "ip": ip, "query_count": 0,
                    "last_seen_ago_sec": 999999, "first_seen_sec_ago": 0,
                    "last_path": "", "ua_short": "",
                    "note": _LAN_DEVICE_NOTES.get(ip, ""),
                    "blocked": True,
                })
        # 按"最后活跃时间"倒序
        devices.sort(key=lambda x: x["last_seen_ago_sec"])
        return jsonify({
            "enabled": _LAN_ENABLED,
            "total_queries": _LAN_STATS["total_requests"],
            "devices": devices,
            "device_count": len(devices),
        })


def _short_ua(ua):
    """把长串 User-Agent 缩成"设备类型 / 浏览器"简写"""
    if not ua:
        return ""
    ua_l = ua.lower()
    device = "电脑"
    if "iphone" in ua_l: device = "iPhone"
    elif "ipad" in ua_l: device = "iPad"
    elif "android" in ua_l: device = "Android"
    elif "macintosh" in ua_l or "mac os x" in ua_l: device = "Mac"
    elif "windows" in ua_l: device = "Windows"
    browser = ""
    if "edg/" in ua_l: browser = "Edge"
    elif "chrome/" in ua_l: browser = "Chrome"
    elif "safari/" in ua_l: browser = "Safari"
    elif "firefox/" in ua_l: browser = "Firefox"
    return f"{device}·{browser}" if browser else device


@app.route("/api/startup", methods=["POST"])
def api_startup():
    """设置或取消开机自启。仅本机管理员可操作。"""
    if not _is_local_request_ip(request.remote_addr or ""):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    data = request.get_json()
    enable = bool(data.get("enable", False))
    ok, msg = _set_startup(enable)
    enabled = _check_startup_enabled()
    return jsonify({"ok": ok, "enabled": enabled, "message": msg})


def _get_startup_plist_path():
    home = os.path.expanduser("~")
    return os.path.join(home, "Library", "LaunchAgents", "com.appfinder.launch.plist")


# Windows 开机自启使用的所有注册表值名（统一后用 AppQueryTool；AppFinder 是遗留）
_WIN_STARTUP_KEYS = ("AppQueryTool", "AppFinder")


def _win_autostart_command():
    """返回注册表里该写的命令：优先 background.vbs（包模式），否则退回到 python app.py"""
    try:
        base = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        base = os.getcwd()
    vbs = os.path.join(base, "background.vbs")
    if os.path.exists(vbs):
        return f'wscript.exe "{vbs}"'
    if getattr(sys, 'frozen', False):
        return f'"{sys.executable}"'
    return f'"{sys.executable}" "{os.path.abspath(__file__)}"'


def _check_startup_enabled():
    if sys.platform == "darwin":
        return os.path.exists(_get_startup_plist_path())
    elif sys.platform == "win32":
        try:
            import winreg
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Run",
                0, winreg.KEY_READ)
            try:
                for name in _WIN_STARTUP_KEYS:
                    try:
                        winreg.QueryValueEx(key, name)
                        return True
                    except FileNotFoundError:
                        continue
            finally:
                winreg.CloseKey(key)
            return False
        except Exception:
            return False
    return False


def _set_startup(enable):
    if sys.platform == "win32":
        try:
            import winreg
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Run",
                0, winreg.KEY_SET_VALUE)
            if enable:
                # 先清掉所有遗留值，避免两个入口同时存在
                for name in _WIN_STARTUP_KEYS:
                    try:
                        winreg.DeleteValue(key, name)
                    except FileNotFoundError:
                        pass
                cmd = _win_autostart_command()
                winreg.SetValueEx(key, "AppQueryTool", 0, winreg.REG_SZ, cmd)
                msg = "已添加开机启动"
            else:
                # 两个值都尝试删（包含遗留的 AppFinder）
                for name in _WIN_STARTUP_KEYS:
                    try:
                        winreg.DeleteValue(key, name)
                    except FileNotFoundError:
                        pass
                msg = "已取消开机启动"
            winreg.CloseKey(key)
            return True, msg
        except Exception as e:
            return False, f"操作失败: {e}"
    if sys.platform != "darwin":
        return False, "开机自启仅支持 Mac 和 Windows"
    plist_path = _get_startup_plist_path()
    if enable:
        # 找到当前可执行文件或 python 脚本
        if getattr(sys, 'frozen', False):
            program = sys.executable
        else:
            program = sys.executable
            script = os.path.abspath(__file__)
            plist_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0">
<dict>
    <key>Label</key>
    <string>com.appfinder.launch</string>
    <key>ProgramArguments</key>
    <array>
        <string>{program}</string>
        <string>{script}</string>
    </array>
    <key>RunAtLoad</key>
    <true/>
    <key>StandardOutPath</key>
    <string>/tmp/appfinder.log</string>
    <key>StandardErrorPath</key>
    <string>/tmp/appfinder.err</string>
</dict>
</plist>"""
        try:
            os.makedirs(os.path.dirname(plist_path), exist_ok=True)
            with open(plist_path, "w") as f:
                f.write(plist_content)
            import subprocess
            subprocess.run(["launchctl", "load", plist_path], capture_output=True)
            return True, "已添加开机启动"
        except Exception as e:
            return False, f"添加失败: {e}"
    else:
        try:
            if os.path.exists(plist_path):
                import subprocess
                subprocess.run(["launchctl", "unload", plist_path], capture_output=True)
                os.remove(plist_path)
            return True, "已取消开机启动"
        except Exception as e:
            return False, f"取消失败: {e}"


@app.route("/api/start_job", methods=["POST"])
def api_start_job():
    """创建后台查询 job，立即返回 job_id（不等查询完成）"""
    _cleanup_jobs()
    req_data = request.get_json()
    tasks, worker_params, meta, est_seconds = _parse_query_input(req_data)
    job_id = uuid.uuid4().hex[:12]
    with JOBS_LOCK:
        JOBS[job_id] = {
            'status':        'running',
            'events':        [],
            'tasks':         tasks,
            'worker_params': worker_params,
            'results':       [],
            'created_at':    time.time(),
        }
    threading.Thread(target=_job_worker, args=(job_id,), daemon=True).start()
    return jsonify({
        'job_id':          job_id,
        'estimated_seconds': est_seconds,
        'total_tasks':     worker_params['total_tasks'],
    })


@app.route("/api/job_stream/<job_id>")
def api_job_stream(job_id):
    """SSE 流：从 offset 开始推送 job 事件，支持断线重连"""
    offset = int(request.args.get('offset', 0))

    def generate():
        idx = offset
        while True:
            job = JOBS.get(job_id)
            if not job:
                yield f"data: {json.dumps({'type': 'error', 'message': 'job expired or not found'})}\n\n"
                return
            events = job['events']
            while idx < len(events):
                yield f"data: {json.dumps(events[idx])}\n\n"
                idx += 1
            if job['status'] != 'running':
                return
            time.sleep(0.2)

    resp = Response(stream_with_context(generate()), mimetype="text/event-stream")
    resp.headers["Cache-Control"]     = "no-cache"
    resp.headers["X-Accel-Buffering"] = "no"
    return resp


@app.route("/api/job_status/<job_id>")
def api_job_status(job_id):
    """查询 job 状态，用于页面重新打开时恢复。内存没有时回落到磁盘缓存"""
    job = JOBS.get(job_id)
    if job:
        return jsonify({
            'found':        True,
            'status':       job['status'],
            'events_count': len(job['events']),
            'results':      job['results'] if job['status'] == 'done' else [],
        })
    # 内存中不存在（服务重启），尝试从磁盘恢复
    disk_results = _load_job_result(job_id)
    if disk_results is not None:
        return jsonify({
            'found':        True,
            'status':       'done',
            'events_count': 0,
            'results':      disk_results,
        })
    return jsonify({'found': False})


@app.route("/api/query", methods=["POST"])
def api_query():
    """查询接口：SSE 流式响应，实时推送进度"""
    req_data = request.get_json()
    raw_inputs = req_data.get("package_names", [])
    exact_search = bool(req_data.get("exact_search", False))
    get_apk_url  = bool(req_data.get("get_apk_url", False))
    apk_url_mode = req_data.get("apk_url_mode", "single")   # "single" | "multiple"
    get_sha1     = bool(req_data.get("get_sha1", False))
    get_sha256   = bool(req_data.get("get_sha256", False))
    platform_filter = req_data.get("platform_filter", "all")  # "all"|"ios"|"android"
    query_interval_ms = max(0, int(req_data.get("query_interval_ms", 0)))


    # ── 输入解析（同步，generator 外完成）──────────────────────────
    MAX_ITEMS = 10000
    pkg_list, ios_id_list, name_list = [], [], []
    invalid_count = 0

    all_valid = [x.strip() for x in raw_inputs if x.strip()]
    total_raw = len(all_valid)
    skipped_over_limit = max(0, total_raw - MAX_ITEMS)

    for item in all_valid[:MAX_ITEMS]:
        if _IOS_ID_RE.match(item):
            ios_id_list.append(item)
        else:
            cleaned = clean_package_name(item)
            if cleaned and is_package_name(cleaned):
                pkg_list.append(cleaned)
            elif re.search(r'[a-zA-Z\u4e00-\u9fff]', item):
                name_list.append(item)
            else:
                invalid_count += 1

    def dedup(lst):
        seen, out = set(), []
        for x in lst:
            if x not in seen:
                seen.add(x); out.append(x)
        return out

    cleaned_pkgs    = dedup(pkg_list)
    cleaned_ios_ids = dedup(ios_id_list)
    cleaned_names   = dedup(name_list)

    total_valid = len(cleaned_pkgs) + len(cleaned_ios_ids) + len(cleaned_names)
    deduplicated = max(0, (total_raw - skipped_over_limit) - total_valid - invalid_count)

    meta = {
        "over_limit":          skipped_over_limit,
        "total_input":         total_raw,
        "invalid_count":       invalid_count,
        "deduplicated":        deduplicated,
        "name_search_ios_only": False,  # 已支持安卓，废弃此字段
    }

    # 全部任务列表（保持输入顺序：包名 → iOS数字ID → 名称）
    tasks = (
        [("pkg",    p) for p in cleaned_pkgs]    +
        [("ios_id", a) for a in cleaned_ios_ids] +
        [("name",   n) for n in cleaned_names]
    )
    total_tasks = len(tasks)

    # ── 批量 / 并发参数 ──────────────────────────────────────────────
    # 用户设置的查询间隔（随机范围 0.8x ~ 1.2x）
    user_interval_s = query_interval_ms / 1000.0 if query_interval_ms > 0 else 0.0

    # 并发上限与 CPU 核数挂钩：每个外层 worker 内部还会再派生 9 个商店子线程
    _CPU_CAP = min((os.cpu_count() or 4), 8)
    if total_tasks <= 500:
        BATCH_SIZE  = total_tasks or 1
        WORKERS     = min(_CPU_CAP, max(1, total_tasks))
        BATCH_DELAY = user_interval_s  # 使用用户设置
        est_seconds = int(math.ceil(total_tasks / max(WORKERS, 1)) * 4 + total_tasks * user_interval_s)
    else:
        BATCH_SIZE  = 20
        WORKERS     = _CPU_CAP
        BATCH_DELAY = max(1.0, user_interval_s)  # 大量查询至少1s间隔
        batches     = math.ceil(total_tasks / BATCH_SIZE)
        est_seconds = int(batches * (7 + BATCH_DELAY))

    # ── SSE Generator ────────────────────────────────────────────────
    def generate():
        if total_tasks == 0:
            yield f"data: {json.dumps({'type': 'complete', 'results': [], **meta})}\n\n"
            return

        yield f"data: {json.dumps({'type': 'start', 'total': total_tasks, 'estimated_seconds': est_seconds, **meta})}\n\n"

        all_results = []
        seen_keys   = set()
        done_count  = 0

        for batch_start in range(0, len(tasks), BATCH_SIZE):
            batch = tasks[batch_start: batch_start + BATCH_SIZE]

            with concurrent.futures.ThreadPoolExecutor(
                    max_workers=min(WORKERS, len(batch))) as exe:
                future_map = {}
                for task_type, value in batch:
                    if task_type == "pkg":
                        if get_apk_url or get_sha1 or get_sha256:
                            f = exe.submit(query_single_extended, value,
                                           None, get_apk_url,
                                           apk_url_mode, get_sha1, get_sha256)
                        else:
                            f = exe.submit(query_single, value)
                    elif task_type == "ios_id":
                        f = exe.submit(search_apple_by_numid, value)
                    else:
                        if get_apk_url or get_sha1 or get_sha256:
                            f = exe.submit(query_by_name_extended, value,
                                           None, exact_search,
                                           get_apk_url, apk_url_mode, get_sha1, get_sha256)
                        else:
                            f = exe.submit(query_by_name, value, None, exact_search)
                    future_map[f] = (task_type, value)

                for f in concurrent.futures.as_completed(future_map):
                    task_type, value = future_map[f]
                    try:
                        result = f.result()
                    except Exception:
                        result = None

                    # 整理本条结果（无结果时保证七麦兜底）
                    rows = _make_fallback_rows(task_type, value, result)

                    new_rows = []
                    for r in rows:
                        # 平台筛选
                        plat = r.get("platform", "")
                        if platform_filter == "ios" and plat != "iOS":
                            continue
                        if platform_filter == "android" and plat != "Android":
                            continue
                        k = (r.get("package_name", ""), plat)
                        if k not in seen_keys:
                            seen_keys.add(k)
                            all_results.append(r)
                            new_rows.append(r)

                    done_count += 1
                    yield f"data: {json.dumps({'type': 'progress', 'done': done_count, 'total': total_tasks, 'rows': new_rows})}\n\n"

            # 批次间限速延迟（加随机抖动 ±20%）
            if BATCH_DELAY > 0 and batch_start + BATCH_SIZE < len(tasks):
                import random
                jitter = BATCH_DELAY * random.uniform(0.8, 1.2)
                time.sleep(jitter)

        yield f"data: {json.dumps({'type': 'complete', 'results': all_results, **meta})}\n\n"

    resp = Response(stream_with_context(generate()), mimetype="text/event-stream")
    resp.headers["Cache-Control"]     = "no-cache"
    resp.headers["X-Accel-Buffering"] = "no"
    return resp


@app.route("/api/download", methods=["POST"])
def api_download():
    """下载结果为xlsx或csv"""
    data = request.get_json()
    results = data.get("results", [])
    file_format = data.get("format", "xlsx")
    include_icon = data.get("include_icon", False)
    include_icon_image = data.get("include_icon_image", False) and file_format == "xlsx" and HAS_PILLOW

    if not results:
        return jsonify({"error": "没有数据可下载"}), 400

    has_apk_urls = any(r.get("apk_direct_urls") for r in results)
    has_sha1     = any(r.get("sha1") for r in results)
    has_sha256   = any(r.get("sha256") for r in results)

    headers_list = ["App名称", "包名", "平台", "分类", "商店地址"]
    if has_apk_urls:
        headers_list.append("下载地址")
    if has_sha1:
        headers_list.append("SHA1")
    if has_sha256:
        headers_list.append("SHA256")

    def apk_urls_str(r):
        urls = r.get("apk_direct_urls", [])
        return "\n".join(urls) if urls else ""

    if file_format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        if include_icon:
            writer.writerow(headers_list + ["图标链接"])
        else:
            writer.writerow(headers_list)
        for r in results:
            row = [r["app_name"], r["package_name"], r["platform"], r.get("category", ""), r["download_url"]]
            if has_apk_urls:
                row.append(apk_urls_str(r))
            if has_sha1:
                row.append(r.get("sha1", ""))
            if has_sha256:
                row.append(r.get("sha256", ""))
            if include_icon:
                row.append(r.get("icon_url", ""))
            writer.writerow(row)
        mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        return send_file(mem, mimetype="text/csv", as_attachment=True, download_name="app_results.csv")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "App查询结果"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # 列定义：含图标图片时第一列为图标列
        col_offset = 0
        if include_icon_image:
            col_offset = 1
            img_header_cell = ws.cell(row=1, column=1, value="图标")
            img_header_cell.font = header_font
            img_header_cell.fill = header_fill
            img_header_cell.alignment = header_alignment
            img_header_cell.border = thin_border
            ws.column_dimensions["A"].width = 6

        cols = headers_list[:]
        if include_icon:
            cols.append("图标链接")

        from openpyxl.utils import get_column_letter
        for col_i, header in enumerate(cols, 1 + col_offset):
            cell = ws.cell(row=1, column=col_i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, r in enumerate(results, 2):
            values = [r["app_name"], r["package_name"], r["platform"], r.get("category", ""), r["download_url"]]
            if has_apk_urls:
                values.append(apk_urls_str(r))
            if has_sha1:
                values.append(r.get("sha1", ""))
            if has_sha256:
                values.append(r.get("sha256", ""))
            if include_icon:
                values.append(r.get("icon_url", ""))
            for col_i, val in enumerate(values, 1 + col_offset):
                cell = ws.cell(row=row_idx, column=col_i, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=bool(has_apk_urls and "\n" in str(val)))

        # 下载并嵌入图标图片
        if include_icon_image:
            def fetch_icon(args):
                row_i, url = args
                if not url:
                    return row_i, None
                try:
                    resp = _HTTP.get(url, timeout=(3, 5), headers={"User-Agent": "Mozilla/5.0"})
                    if resp.status_code == 200:
                        img = PILImage.open(io.BytesIO(resp.content)).convert("RGBA")
                        img.thumbnail((36, 36), PILImage.LANCZOS)
                        buf = io.BytesIO()
                        img.save(buf, format="PNG")
                        buf.seek(0)
                        return row_i, buf
                except Exception:
                    pass
                return row_i, None

            tasks = [(i, results[i - 2].get("icon_url", "")) for i in range(2, len(results) + 2)]
            with concurrent.futures.ThreadPoolExecutor(max_workers=20) as exe:
                for row_i, buf in exe.map(fetch_icon, tasks):
                    if buf:
                        try:
                            xl_img = XLImage(buf)
                            xl_img.width = 36
                            xl_img.height = 36
                            ws.add_image(xl_img, f"A{row_i}")
                            ws.row_dimensions[row_i].height = 28
                        except Exception:
                            pass

        base_widths = [20, 35, 12, 15, 55]
        extra_widths = []
        if has_apk_urls:
            extra_widths.append(70)
        if has_sha1:
            extra_widths.append(55)
        if include_icon:
            extra_widths.append(50)
        all_widths = base_widths + extra_widths

        start_col = (2 if include_icon_image else 1)
        for idx, w in enumerate(all_widths):
            letter = get_column_letter(start_col + idx)
            ws.column_dimensions[letter].width = w

        mem = io.BytesIO()
        wb.save(mem)
        mem.seek(0)
        return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="app_results.xlsx")


def _is_real_lan_ip(ip):
    """真正的家庭/办公室 LAN 的 RFC1918 私有地址。
    明确排除：
    - 127.x（回环）/ 169.254.x（link-local）
    - 198.18.x/198.19.x（Cloudflare Warp / 网络基准测试网段）
    - 100.64-127.x（运营商 CGN / Tailscale / Zerotier）
    """
    if not ip:
        return False
    if ip.startswith("192.168."):
        return True
    if ip.startswith("10."):
        return True
    if ip.startswith("172."):
        try:
            if 16 <= int(ip.split(".")[1]) <= 31:
                return True
        except (ValueError, IndexError):
            pass
    return False


def _get_lan_ip():
    """探测本机的**真实 LAN IP**，绕开 VPN/Warp 等改写默认路由的工具。
    策略（按可靠度降序）：
      1. macOS: `ipconfig getifaddr enN` 遍历物理网卡——直接读取网卡级 IP，
         不受 VPN 默认路由影响
      2. socket.gethostbyname_ex(hostname) 枚举所有本机 IP，筛私有网段
      3. connect 8.8.8.8 UDP（原逻辑，兜底）
    只返回真正的 RFC1918 私有地址；全部失败返回 127.0.0.1。
    """
    import socket
    import subprocess

    # Method 1: Mac 下直接读物理网卡 IP，绕过 VPN
    if sys.platform == "darwin":
        # en0 en1 ... en9，Mac 标准命名
        for iface in ["en0", "en1", "en2", "en3", "en4", "en5"]:
            try:
                r = subprocess.run(
                    ["ipconfig", "getifaddr", iface],
                    capture_output=True, text=True, timeout=1.5,
                )
                ip = (r.stdout or "").strip()
                if _is_real_lan_ip(ip):
                    return ip
            except Exception:
                continue

    # Method 2: 枚举本机所有 IP，挑 RFC1918 的
    try:
        _, _, ips = socket.gethostbyname_ex(socket.gethostname())
        private = [ip for ip in ips if _is_real_lan_ip(ip)]
        if private:
            # 优先 192.168.x（家用路由器最常见），其次 10.x，再 172.16-31.x
            private.sort(key=lambda x: (
                0 if x.startswith("192.168.") else
                1 if x.startswith("10.") else 2
            ))
            return private[0]
    except Exception:
        pass

    # Method 3: UDP connect 探测（兜底，可能被 VPN 干扰）
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        if _is_real_lan_ip(ip):
            return ip
    except Exception:
        pass
    finally:
        s.close()

    return "127.0.0.1"


if __name__ == "__main__":
    import webbrowser
    import threading

    # 启动时加载持久化的 LAN 开关设置 + 共享历史
    _load_lan_settings()
    _load_history()

    port = int(os.environ.get("PORT", 9527))

    try:
        lan_ip = _get_lan_ip()
        url = f"http://{lan_ip}:{port}"
        # 自动打开浏览器的决策：
        # - Mac .app 打包（sys.frozen + darwin）：没有 launcher 代为打开，必须自己开
        # - Windows：launch.py 会开浏览器，这里跳过
        # - 源码直跑（开发模式）：不开，不然每次 reload 都弹
        # - APPFINDER_NO_BROWSER=1：任何情况都不开（给包装方式覆盖）
        is_frozen = getattr(sys, 'frozen', False)
        is_mac_bundle = is_frozen and sys.platform == 'darwin'
        want_open = is_mac_bundle and not os.environ.get("APPFINDER_NO_BROWSER")

        print(f"\n  App Query Tool Started!")
        print(f"  Open in browser:  {url}")
        print(f"  Also accessible:  http://127.0.0.1:{port}  (this machine only)")
        print(f"  Close this window to exit.\n")

        if is_mac_bundle:
            # Mac .app 特殊处理：
            #   - Flask 在后台 daemon 线程跑（主进程退出它自动结束）
            #   - 主线程跑一个隐藏的 tkinter 主循环，让 macOS 把它识别为图形应用
            #     从而 Dock 图标持续可见、能响应 Cmd-Q / 右键 → 退出
            # 浏览器自动打开用 127.0.0.1 而非 LAN IP：
            # - 127.0.0.1 走回环，网络隔离/VPN/DNS 都不会干扰
            # - LAN 开关关闭时也能访问（门禁只拦非本机请求）
            # - URL 栏显示 127.0.0.1 也更清晰："这是自己电脑"
            if want_open:
                local_url = f"http://127.0.0.1:{port}"
                threading.Timer(1.5, lambda: webbrowser.open(local_url)).start()

            flask_thread = threading.Thread(
                target=lambda: app.run(host='0.0.0.0', port=port,
                                       debug=False, use_reloader=False),
                daemon=True,
            )
            flask_thread.start()

            import tkinter as tk
            root = tk.Tk()
            root.withdraw()  # 窗口不可见，但 NSApplication 已注册，Dock 图标留存
            root.title("App 查询工具")
            # Cmd-Q 或 Dock 右键"退出"：关闭 Tk 主循环 → 主进程退 → Flask 守护线程随之退出
            try:
                root.createcommand('tk::mac::Quit', root.quit)
            except Exception:
                pass
            root.protocol("WM_DELETE_WINDOW", root.quit)
            root.mainloop()
        else:
            # Windows 或开发模式：原来的阻塞式启动
            # 浏览器自动打开用 127.0.0.1 而非 LAN IP：
            # - 127.0.0.1 走回环，网络隔离/VPN/DNS 都不会干扰
            # - LAN 开关关闭时也能访问（门禁只拦非本机请求）
            # - URL 栏显示 127.0.0.1 也更清晰："这是自己电脑"
            if want_open:
                local_url = f"http://127.0.0.1:{port}"
                threading.Timer(1.5, lambda: webbrowser.open(local_url)).start()
            # 绑 0.0.0.0：本机、局域网其他设备、以及某些严格防火墙下都能访问
            app.run(host='0.0.0.0', port=port, debug=False)
    except Exception as e:
        print(f"\n  ERROR: {e}")
        input("\n  Press Enter to exit...")
